import io
import os
import uuid

import pytest
import requests

BASE_URL = os.environ.get("BACKEND_TEST_URL", "http://localhost:8001").rstrip("/")
API = f"{BASE_URL}/api"

ADMIN_EMAIL = "admin@plutusventures.com"
ADMIN_PASSWORD = "admin123"

# ---------------------------- fixtures ----------------------------

@pytest.fixture(scope="session")
def admin_token():
    r = requests.post(
        f"{API}/auth/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
        timeout=10,
    )
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    data = r.json()
    assert "token" in data, f"login response missing token: {data}"
    return data["token"]

@pytest.fixture(scope="session")
def auth(admin_token):
    s = requests.Session()
    s.headers.update({
        "Authorization": f"Bearer {admin_token}",
        "Content-Type": "application/json",
    })
    return s

@pytest.fixture(scope="session")
def company_id(auth):
    """Create a dedicated test company for this run."""
    name = f"TEST_HistCo_{uuid.uuid4().hex[:8]}"
    payload = {
        "company_name": name,
        "contact_person": "QA Bot",
        "phone": "9999999999",
        "email": "qa@example.com",
        "client_email": "client@example.com",
        "address": "1 Test Rd",
        "city": "Bengaluru",
        "state": "KA",
        "pincode": "560001",
    }
    r = auth.post(f"{API}/companies", json=payload)
    assert r.status_code in (200, 201), f"company create failed: {r.status_code} {r.text}"
    cid = r.json().get("id")
    assert cid, f"no id in company response: {r.json()}"
    return cid

@pytest.fixture
def sub_admin_auth(auth):
    """Create a sub-admin with no company assignments and return an auth session."""
    email = f"subadmin_{uuid.uuid4().hex[:8]}@example.com"
    password = "pass1234"
    r = auth.post(f"{API}/sub-admins", json={
        "name": "QA Sub Admin",
        "email": email,
        "password": password,
    })
    assert r.status_code in (200, 201), f"sub-admin create failed: {r.status_code} {r.text}"
    assert r.json().get("assigned_company_ids") in ([], None)

    login = requests.post(
        f"{API}/auth/login",
        json={"email": email, "password": password},
        timeout=10,
    )
    assert login.status_code == 200, f"sub-admin login failed: {login.status_code} {login.text}"
    token = login.json().get("token")
    assert token, f"sub-admin login response missing token: {login.json()}"
    s = requests.Session()
    s.headers.update({
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    })
    return s

@pytest.fixture
def ticket_admin_auth(auth):
    """Create a limited ticket admin and return an auth session."""
    email = f"ticketadmin_{uuid.uuid4().hex[:8]}@example.com"
    password = "pass1234"
    r = auth.post(f"{API}/ticket-admins", json={
        "name": "QA Ticket Admin",
        "email": email,
        "password": password,
    })
    assert r.status_code in (200, 201), f"ticket-admin create failed: {r.status_code} {r.text}"
    assert r.json().get("role") == "ticket_admin"

    login = requests.post(
        f"{API}/auth/login",
        json={"email": email, "password": password},
        timeout=10,
    )
    assert login.status_code == 200, f"ticket-admin login failed: {login.status_code} {login.text}"
    token = login.json().get("token")
    assert token, f"ticket-admin login response missing token: {login.json()}"
    s = requests.Session()
    s.headers.update({
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    })
    return s

def _new_ticket_payload(company_id_val, *, customer_email="cust@example.com",
                        prod_ref="PRN-001", oem_ref="OEM-001",
                        omit_email=False, bad_email=False):
    payload = {
        "company_id": company_id_val,
        "customer_name": "Test Customer",
        "customer_phone": "9000000000",
        "contact_source": "email",
        "issue_description": "Device not booting",
        "priority": "medium",
        "product_reference_number": prod_ref,
        "oem_reference_number": oem_ref,
        "device": {
            "brand": "Acme",
            "model": "X1",
            "serial_number": f"SN-{uuid.uuid4().hex[:6]}",
            "warranty_status": "active",
        },
    }
    if not omit_email:
        payload["customer_email"] = "not-an-email" if bad_email else customer_email
    return payload

# ---------------------------- Auth basic ----------------------------

class TestAuth:
    def test_login_returns_token_no_otp(self):
        r = requests.post(
            f"{API}/auth/login",
            json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
            timeout=10,
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert "token" in body and isinstance(body["token"], str) and body["token"]
        assert body.get("user", {}).get("email") == ADMIN_EMAIL
        assert body.get("user", {}).get("role") == "admin"

    def test_auth_me(self, auth):
        r = auth.get(f"{API}/auth/me")
        assert r.status_code == 200, r.text
        body = r.json()
        assert body.get("email") == ADMIN_EMAIL

# ---------------------------- Feature 1: Ticket fields ----------------------------

class TestTicketCreation:
    def test_create_ticket_with_new_fields(self, auth, company_id):
        payload = _new_ticket_payload(company_id, prod_ref="PRN-ALPHA",
                                      oem_ref="OEM-BETA")
        r = auth.post(f"{API}/tickets", json=payload)
        assert r.status_code in (200, 201), r.text
        t = r.json()
        assert t["product_reference_number"] == "PRN-ALPHA"
        assert t["oem_reference_number"] == "OEM-BETA"
        assert t["customer_email"] == "cust@example.com"
        # Persistence check via GET
        gid = t["id"]
        g = auth.get(f"{API}/tickets/{gid}")
        assert g.status_code == 200, g.text
        gt = g.json()
        assert gt["product_reference_number"] == "PRN-ALPHA"
        assert gt["oem_reference_number"] == "OEM-BETA"
        assert gt["customer_email"] == "cust@example.com"

    def test_create_ticket_missing_email_422(self, auth, company_id):
        payload = _new_ticket_payload(company_id, omit_email=True)
        r = auth.post(f"{API}/tickets", json=payload)
        assert r.status_code == 422, f"expected 422, got {r.status_code}: {r.text}"

    def test_create_ticket_invalid_email_422(self, auth, company_id):
        payload = _new_ticket_payload(company_id, bad_email=True)
        r = auth.post(f"{API}/tickets", json=payload)
        assert r.status_code == 422, f"expected 422, got {r.status_code}: {r.text}"

    def test_create_ticket_without_ref_numbers_still_works(self, auth, company_id):
        payload = _new_ticket_payload(company_id, prod_ref=None, oem_ref=None)
        # remove None keys to simulate legacy payload
        payload.pop("product_reference_number")
        payload.pop("oem_reference_number")
        r = auth.post(f"{API}/tickets", json=payload)
        assert r.status_code in (200, 201), r.text
        t = r.json()
        assert t.get("product_reference_number") in (None, "")
        assert t.get("oem_reference_number") in (None, "")
        assert t["customer_email"] == "cust@example.com"

    def test_create_ticket_with_multiple_devices(self, auth, company_id):
        serial_a = f"SN-{uuid.uuid4().hex[:6]}"
        serial_b = f"SN-{uuid.uuid4().hex[:6]}"
        payload = _new_ticket_payload(company_id)
        payload.pop("device")
        payload["devices"] = [
            {
                "brand": "Acme",
                "model": "X1",
                "serial_number": serial_a,
                "warranty_status": "active",
            },
            {
                "brand": "Globex",
                "model": "Printer 200",
                "serial_number": serial_b,
                "device_type": "printer",
                "warranty_status": "none",
            },
        ]
        r = auth.post(f"{API}/tickets", json=payload)
        assert r.status_code in (200, 201), r.text
        t = r.json()
        assert t["device_count"] == 2
        assert len(t["device_ids"]) == 2
        assert len(t["devices"]) == 2
        assert t["device_id"] == t["device_ids"][0]
        assert {d["serial_number"] for d in t["devices"]} == {serial_a, serial_b}

        g = auth.get(f"{API}/tickets/{t['id']}")
        assert g.status_code == 200, g.text
        gt = g.json()
        assert gt["device_count"] == 2
        assert len(gt["devices"]) == 2
        assert {d["model"] for d in gt["devices"]} == {"X1", "Printer 200"}

    def test_sub_admin_can_create_ticket_for_any_active_company(self, sub_admin_auth, company_id):
        payload = _new_ticket_payload(company_id, customer_email="subadmin@example.com")
        r = sub_admin_auth.post(f"{API}/tickets", json=payload)
        assert r.status_code in (200, 201), r.text
        t = r.json()
        assert t["company_id"] == company_id
        assert t["customer_email"] == "subadmin@example.com"

    def test_ticket_admin_limited_ticket_permissions(self, auth, ticket_admin_auth, company_id):
        payload = _new_ticket_payload(company_id, customer_email="ticketadmin@example.com")
        r = ticket_admin_auth.post(f"{API}/tickets", json=payload)
        assert r.status_code in (200, 201), r.text
        ticket = r.json()

        engineer_email = f"limited_eng_{uuid.uuid4().hex[:8]}@example.com"
        eng = auth.post(f"{API}/engineers", json={
            "name": "Limited Role Engineer",
            "email": engineer_email,
            "password": "pass1234",
            "skills": [],
        })
        assert eng.status_code in (200, 201), eng.text

        assign = ticket_admin_auth.post(
            f"{API}/tickets/{ticket['id']}/assign",
            json={"engineer_id": eng.json()["id"], "is_outsource": False},
        )
        assert assign.status_code == 200, assign.text
        assert assign.json()["status"] == "assigned"

        outsource = ticket_admin_auth.post(
            f"{API}/tickets/{ticket['id']}/assign",
            json={"is_outsource": True, "outsource_name": "Outside Partner"},
        )
        assert outsource.status_code == 403, outsource.text

        approve = ticket_admin_auth.post(f"{API}/tickets/{ticket['id']}/approve")
        assert approve.status_code == 403, approve.text

        history = ticket_admin_auth.get(f"{API}/device-history")
        assert history.status_code == 200, history.text

# ---------------------------- Feature 2: Approve flow ----------------------------

class TestApproveFlow:
    def test_approve_closes_ticket_and_mocks_email(self, auth, company_id):
        # Create a ticket
        payload = _new_ticket_payload(
            company_id, customer_email="closeme@example.com",
            prod_ref="PRN-CLOSE", oem_ref="OEM-CLOSE",
        )
        r = auth.post(f"{API}/tickets", json=payload)
        assert r.status_code in (200, 201), r.text
        tid = r.json()["id"]

        # Approve as admin
        r2 = auth.post(f"{API}/tickets/{tid}/approve")
        assert r2.status_code == 200, r2.text
        approved = r2.json()
        assert approved["status"] == "closed"
        assert approved.get("approved") is True
        assert approved.get("approved_at"), "approved_at not set"
        assert approved.get("closure_email_sent_to") == "closeme@example.com"
        status = approved.get("closure_email_status")
        assert isinstance(status, dict), f"closure_email_status not dict: {status}"
        # Iter 2: SMTP placeholder Gmail creds may be configured. Approve must NOT
        # crash. Either mocked (no SMTP) or sent=false with error (placeholder creds).
        assert status.get("sent") is False, f"expected sent=false: {status}"
        assert status.get("mocked") is True or "error" in status, (
            f"expected mocked=true OR error key on graceful SMTP failure: {status}"
        )

        # Persistence
        g = auth.get(f"{API}/tickets/{tid}")
        assert g.status_code == 200
        assert g.json()["status"] == "closed"

# ---------------------------- Feature 3: Device history ----------------------------

class TestDeviceHistory:
    @pytest.fixture(scope="class")
    def closed_ticket(self, auth, company_id):
        """Create + approve a ticket so it's available in device-history."""
        payload = _new_ticket_payload(
            company_id, customer_email="hist@example.com",
            prod_ref="PRN-HIST", oem_ref="OEM-HIST",
        )
        r = auth.post(f"{API}/tickets", json=payload)
        assert r.status_code in (200, 201), r.text
        t = r.json()
        a = auth.post(f"{API}/tickets/{t['id']}/approve")
        assert a.status_code == 200, a.text
        return a.json()

    def test_list_device_history(self, auth, closed_ticket):
        r = auth.get(f"{API}/device-history")
        assert r.status_code == 200, r.text
        body = r.json()
        assert "items" in body and "total" in body
        items = body["items"]
        # find our ticket
        match = next((i for i in items if i.get("id") == closed_ticket["id"]), None)
        assert match is not None, "newly closed ticket not present in device-history"
        # required row fields
        required = {"device_id", "ticket_id", "company_name", "engineer_name",
                    "status", "created_date", "closed_date",
                    "product_reference_number", "oem_reference_number"}
        missing = required - set(match.keys())
        assert not missing, f"missing fields in history row: {missing}"
        assert match["status"] == "closed"
        assert match["product_reference_number"] == "PRN-HIST"
        assert match["oem_reference_number"] == "OEM-HIST"
        assert match["closed_date"], "closed_date empty for closed ticket"

    def test_filter_device_history(self, auth, closed_ticket):
        r = auth.get(f"{API}/device-history/filter",
                     params={"company": "TEST_HistCo"})
        assert r.status_code == 200, r.text
        body = r.json()
        assert "items" in body and "total" in body
        ids = {i.get("id") for i in body["items"]}
        assert closed_ticket["id"] in ids

    def test_export_device_history_xlsx(self, auth, closed_ticket):
        params = {"company": "TEST_HistCo",
                  "start_date": "2020-01-01", "end_date": "2099-12-31"}
        r = auth.get(f"{API}/device-history/export", params=params)
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "")
        assert "spreadsheet" in ct, f"unexpected content-type: {ct}"
        cd = r.headers.get("content-disposition", "")
        # filename pattern device_history_<company>_<start>_<end>.xlsx
        assert "device_history_" in cd and ".xlsx" in cd, f"bad cd: {cd}"
        assert "TEST_HistCo" in cd, f"company missing in filename: {cd}"
        # Server slugifies dashes in dates to underscores - accept either
        assert ("2020-01-01" in cd or "2020_01_01" in cd), f"start_date missing: {cd}"
        assert ("2099-12-31" in cd or "2099_12_31" in cd), f"end_date missing: {cd}"

        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers_row = [c.value for c in ws[1]]
        expected = [
            "Device ID", "Ticket ID", "Company Name", "Engineer Name", "Status",
            "Created Date", "Closed Date",
            "Product Reference Number", "OEM Reference Number",
        ]
        assert headers_row == expected, f"headers mismatch: {headers_row}"
        # find row matching our ticket
        ticket_no = closed_ticket.get("ticket_no") or closed_ticket.get("ticket_number")
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == ticket_no:
                found = True
                assert row[4] == "closed"
                assert row[7] == "PRN-HIST"
                assert row[8] == "OEM-HIST"
                break
        assert found, f"ticket {ticket_no} not found in exported xlsx"

    def test_delete_open_ticket_400(self, auth, company_id):
        payload = _new_ticket_payload(
            company_id, customer_email="open@example.com",
        )
        r = auth.post(f"{API}/tickets", json=payload)
        assert r.status_code in (200, 201), r.text
        tid = r.json()["id"]
        d = auth.delete(f"{API}/device-history/{tid}")
        assert d.status_code == 400, f"expected 400 for open ticket, got {d.status_code}: {d.text}"

    def test_only_deleted_and_include_deleted_filters(self, auth, company_id):
        """Iter 2: GET /api/device-history/filter supports include_deleted/only_deleted."""
        # Create + approve + soft-delete a ticket
        payload = _new_ticket_payload(
            company_id, customer_email="todel@example.com",
            prod_ref="PRN-DEL", oem_ref="OEM-DEL",
        )
        r = auth.post(f"{API}/tickets", json=payload)
        assert r.status_code in (200, 201), r.text
        tid = r.json()["id"]
        a = auth.post(f"{API}/tickets/{tid}/approve")
        assert a.status_code == 200
        d = auth.delete(f"{API}/device-history/{tid}")
        assert d.status_code == 200

        # Default: soft-deleted excluded
        r1 = auth.get(f"{API}/device-history/filter")
        ids_default = {i.get("id") for i in r1.json().get("items", [])}
        assert tid not in ids_default, "default filter should hide soft-deleted"

        # only_deleted=true: only soft-deleted rows
        r2 = auth.get(f"{API}/device-history/filter", params={"only_deleted": "true"})
        assert r2.status_code == 200, r2.text
        items2 = r2.json().get("items", [])
        ids_only = {i.get("id") for i in items2}
        assert tid in ids_only, "only_deleted=true should surface soft-deleted ticket"
        # every row must be is_deleted=true
        for i in items2:
            assert i.get("is_deleted") is True, f"non-deleted leaked into only_deleted: {i.get('id')}"

        # include_deleted=true: both
        r3 = auth.get(f"{API}/device-history/filter",
                      params={"include_deleted": "true"})
        assert r3.status_code == 200, r3.text
        ids_inc = {i.get("id") for i in r3.json().get("items", [])}
        assert tid in ids_inc, "include_deleted=true should contain soft-deleted ticket"

        # Cleanup: restore so other tests aren't affected
        auth.post(f"{API}/device-history/{tid}/restore")

    def test_soft_delete_and_restore(self, auth, closed_ticket):
        tid = closed_ticket["id"]
        # Soft-delete
        d = auth.delete(f"{API}/device-history/{tid}")
        assert d.status_code == 200, d.text
        body = d.json()
        assert body.get("is_deleted") is True

        # Excluded from device-history
        r = auth.get(f"{API}/device-history")
        ids = {i.get("id") for i in r.json().get("items", [])}
        assert tid not in ids, "soft-deleted ticket still in device-history"

        # Excluded from /api/tickets
        r2 = auth.get(f"{API}/tickets")
        assert r2.status_code == 200
        ids2 = {t.get("id") for t in r2.json()}
        assert tid not in ids2, "soft-deleted ticket still in /api/tickets"

        # Restore
        rest = auth.post(f"{API}/device-history/{tid}/restore")
        assert rest.status_code == 200, rest.text
        assert rest.json().get("is_deleted") is False

        # Re-appears
        r3 = auth.get(f"{API}/device-history")
        ids3 = {i.get("id") for i in r3.json().get("items", [])}
        assert tid in ids3, "restored ticket missing from device-history"

# ---------------------------- Backward-compat endpoints ----------------------------

class TestBackwardCompat:
    def test_companies_list(self, auth):
        r = auth.get(f"{API}/companies")
        assert r.status_code == 200, r.text
        body = r.json()
        # Endpoint supports both list and paginated dict formats
        if isinstance(body, dict):
            assert "items" in body and isinstance(body["items"], list)
        else:
            assert isinstance(body, list)

    def test_dashboard_admin(self, auth):
        r = auth.get(f"{API}/dashboard/admin")
        assert r.status_code == 200, r.text
        assert isinstance(r.json(), dict)

    def test_legacy_history_export(self, auth):
        r = auth.get(f"{API}/devices/history-export")
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "")
        assert "spreadsheet" in ct, f"unexpected content-type: {ct}"
