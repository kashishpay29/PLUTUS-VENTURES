from dotenv import load_dotenv
import os
import certifi
load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

import os
import uuid
import logging
import random
import time
from datetime import datetime, timezone, timedelta
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import date
from pathlib import Path
from typing import List, Optional, Literal, Dict, Any

from fastapi import (
    FastAPI, APIRouter, HTTPException, Depends, Request, Response,
    Query, UploadFile, File
)
import io
import openpyxl

from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.middleware.gzip import GZipMiddleware
from pymongo import MongoClient
from pymongo.errors import DuplicateKeyError
from pydantic import BaseModel, EmailStr
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

from auth import (
    hash_password, verify_password, create_access_token,
    get_current_user, require_admin, require_engineer, require_sub_admin,
    require_admin_console, require_ticket_operator, require_company_admin, decode_token,
)
from storage_client import init_storage, put_object, get_object
from pdf_gen import build_service_report_pdf, build_outsource_internal_pdf

# ---------- Setup ----------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(name)s | %(message)s'
)
logger = logging.getLogger("plutus-serviceops")

# ---------- Firebase ----------
import firebase_admin
from firebase_admin import credentials, messaging as fcm_messaging

def _init_firebase():
    try:
        firebase_admin.initialize_app(options={
            "projectId": "plutus-ventures"
        })
        logger.info("✓ Firebase initialized")
    except Exception as e:
        logger.error(f"Firebase init failed: {e}")

_init_firebase()

def send_push_notification(token: str, title: str, body: str, data: dict = None):
    try:
        message = fcm_messaging.Message(
            notification=fcm_messaging.Notification(title=title, body=body),
            data={k: str(v) for k, v in (data or {}).items()},
            token=token,
        )
        fcm_messaging.send(message)
        logger.info(f"Push sent: {title}")
    except Exception as e:
        logger.error(f"Push notification failed: {e}")
_RESPONSE_CACHE: Dict[str, Any] = {}

def cache_key(prefix: str, **kwargs) -> str:
    return f"{prefix}:{':'.join(f'{k}={v}' for k, v in sorted(kwargs.items()))}"

def get_cache(key: str, max_age_seconds: int = 300) -> Optional[Any]:
    if key in _RESPONSE_CACHE:
        data, timestamp = _RESPONSE_CACHE[key]
        if time.time() - timestamp < max_age_seconds:
            return data
        del _RESPONSE_CACHE[key]
    return None

def set_cache(key: str, data: Any):
    _RESPONSE_CACHE[key] = (data, time.time())


def send_ticket_email(to_email: str, ticket: dict, pdf_bytes: bytes = None,
                      subject_prefix: str = "", sender_name: str = None, sender_email: str = None):
    """Send ticket notification email with optional PDF attachment."""
    try:
        smtp_host = os.environ.get("SMTP_HOST", "")
        smtp_port = int(os.environ.get("SMTP_PORT", "587"))
        smtp_user = os.environ.get("SMTP_USER", "")
        smtp_pass = os.environ.get("SMTP_PASS", "")

        if not smtp_host or not smtp_user:
            logger.warning("SMTP not configured - skipping email")
            return

        # Use sender's name/email if provided, else fall back to env
        display_name = sender_name or os.environ.get("COMPANY_NAME", "Plutus Ventures")
        reply_to = sender_email or os.environ.get("FROM_EMAIL", smtp_user)
        from_email = f"{display_name} <{smtp_user}>"

        ticket_no = ticket.get("ticket_no") or ticket.get("ticket_number", "")
        subject = f"{subject_prefix}Service Ticket {ticket_no} - Plutus Ventures"

        msg = MIMEMultipart()
        msg["From"] = from_email
        msg["Reply-To"] = reply_to
        msg["To"] = to_email
        msg["Subject"] = subject

        status = ticket.get("status", "").replace("_", " ").title()
        device_label = ", ".join(_ticket_device_ids(ticket)) or ticket.get("device_id", "—")
        body = f"""
Dear Customer,
Your service ticket <b>{ticket_no}</b> has been <b>{status}</b>.
<b>Details:</b>
- Ticket: {ticket_no}
- Status: {status}
- Device: {device_label}
- Issue: {ticket.get("issue_description") or ticket.get("problem_description", "—")}
{"Please find the service report attached." if pdf_bytes else ""}
Thank you for choosing Plutus Ventures.
Regards,
Plutus Ventures IT Service Team
        """.strip()

        msg.attach(MIMEText(body, "html"))

        if pdf_bytes:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(pdf_bytes)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename=ServiceReport_{ticket_no}.pdf")
            msg.attach(part)

        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(from_email, to_email, msg.as_string())

        logger.info(f"Email sent to {to_email} for ticket {ticket_no}")
    except Exception as e:
        logger.error(f"Email send failed: {e}")

def send_ticket_closed_email(to_email: str, ticket: dict, pdf_bytes: bytes = None,
                             sender_name: str = None, sender_email: str = None):
    """Send 'Ticket Closed' email with engineer-submitted PDF attached.

    Subject: Ticket Closed - {ticket_id}
    Body includes Ticket ID, Product Reference Number, OEM Reference Number,
    Approval Date & Time and a closure confirmation message.
    """
    try:
        smtp_host = os.environ.get("SMTP_HOST", "")
        smtp_port = int(os.environ.get("SMTP_PORT", "587"))
        smtp_user = os.environ.get("SMTP_USER", "")
        smtp_pass = os.environ.get("SMTP_PASS", "")

        ticket_no = ticket.get("ticket_no") or ticket.get("ticket_number", "")
        approved_at = ticket.get("approved_at") or now_iso()
        # Pretty format approved_at (YYYY-MM-DD HH:MM UTC) if ISO
        try:
            dt = datetime.fromisoformat(approved_at.replace("Z", "+00:00"))
            approved_at_pretty = dt.strftime("%Y-%m-%d %H:%M UTC")
        except Exception:
            approved_at_pretty = approved_at

        product_ref = ticket.get("product_reference_number") or "—"
        oem_ref = ticket.get("oem_reference_number") or "—"

        subject = f"Ticket Closed - {ticket_no}"
        display_name = sender_name or os.environ.get("COMPANY_NAME", "Plutus Ventures")
        reply_to = sender_email or os.environ.get("FROM_EMAIL", smtp_user or "no-reply@plutusventures.com")

        body = f"""
Dear Customer,
Your service ticket <b>{ticket_no}</b> has been <b>closed</b> after successful resolution and approval.
<b>Ticket Details:</b><br/>
- Ticket ID: {ticket_no}<br/>
- Product Reference Number: {product_ref}<br/>
- OEM Reference Number: {oem_ref}<br/>
- Approval Date &amp; Time: {approved_at_pretty}<br/>
The engineer's service report is attached for your records. If you have any
questions or need further assistance, please reply to this email.
Thank you for choosing {display_name}.
Regards,<br/>
{display_name} IT Service Team
        """.strip()

        # If SMTP not configured, log the email and return (mocked mode for dev/tests).
        if not smtp_host or not smtp_user:
            logger.warning(
                f"SMTP not configured - MOCK email | to={to_email} subject='{subject}' "
                f"pdf_attached={bool(pdf_bytes)}"
            )
            return {"sent": False, "mocked": True, "to": to_email, "subject": subject}

        from_email = f"{display_name} <{smtp_user}>"
        msg = MIMEMultipart()
        msg["From"] = from_email
        msg["Reply-To"] = reply_to
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "html"))

        if pdf_bytes:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(pdf_bytes)
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename=ServiceReport_{ticket_no}.pdf",
            )
            msg.attach(part)

        # Simple retry: try up to 2 attempts
        last_err = None
        for attempt in range(2):
            try:
                with smtplib.SMTP(smtp_host, smtp_port) as server:
                    server.starttls()
                    server.login(smtp_user, smtp_pass)
                    server.sendmail(from_email, to_email, msg.as_string())
                logger.info(f"Closure email sent to {to_email} for ticket {ticket_no}")
                return {"sent": True, "to": to_email, "subject": subject}
            except Exception as e:
                last_err = e
                logger.error(f"Closure email attempt {attempt + 1} failed: {e}")
        return {"sent": False, "error": str(last_err), "to": to_email, "subject": subject}
    except Exception as e:
        logger.error(f"send_ticket_closed_email failed: {e}")
        return {"sent": False, "error": str(e), "to": to_email}

mongo_url = os.environ["MONGO_URL"]

import ssl

client = MongoClient(
    mongo_url,
    tlsAllowInvalidCertificates=True,
    maxPoolSize=50,
    minPoolSize=10,
    socketTimeoutMS=10000,
    connectTimeoutMS=10000,
    serverSelectionTimeoutMS=5000,
    maxIdleTimeMS=45000,
    waitQueueTimeoutMS=10000,
    retryWrites=True
)

db = client[os.environ["DB_NAME"]]

# Rate limiter
limiter = Limiter(key_func=get_remote_address)

app = FastAPI(
    title="Plutus Ventures – IT Service Management API",
    version="2.1.0",
    docs_url=os.environ.get("DOCS_URL", "/api/docs"),
    redoc_url=None,
)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
api = APIRouter(prefix="/api")

# ---------- Security headers + request logging ----------
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "camera=(self), geolocation=(self)"
        if request.url.scheme == "https":
            response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        return response

class RequestLogMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        import time
        start = time.time()
        response = await call_next(request)
        if request.url.path.startswith("/api/"):
            dur = int((time.time() - start) * 1000)
            logger.info(
                f"{request.method} {request.url.path} "
                f"-> {response.status_code} {dur}ms"
            )
        return response

# ---------- Constants ----------
TICKET_STATUSES = [
    "open", "assigned", "accepted", "travelling", "reached_site",
    "in_progress", "resolved",
    "completed_with_signature", "report_generated", "closed",
    "rejected",
]

# ---------- Helpers ----------
def now_iso():
    return datetime.now(timezone.utc).isoformat()


def new_id():
    return str(uuid.uuid4())

def clean(d):
    if not d:
        return d
    d.pop("_id", None)
    d.pop("password_hash", None)
    return d

def _sub_admin_ticket_scope(user: dict) -> Dict[str, Any]:
    """Limit sub-admin ticket visibility to assigned companies plus own tickets."""
    if user.get("role") != "sub_admin":
        return {}
    assigned_company_ids = user.get("assigned_company_ids") or []
    scope = [{"created_by": user["id"]}]
    if assigned_company_ids:
        scope.append({"company_id": {"$in": assigned_company_ids}})
    return {"$or": scope}

def _cache_key(prefix: str, user: dict, *parts: Any) -> str:
    assigned = ",".join(sorted(user.get("assigned_company_ids") or []))
    values = [prefix, user.get("id", ""), user.get("role", ""), assigned]
    values.extend(str(part) for part in parts)
    return "|".join(values)

def _cache_get(key: str):
    item = _RESPONSE_CACHE.get(key)
    if not item:
        return None
    expires_at, value = item
    if expires_at < time.time():
        _RESPONSE_CACHE.pop(key, None)
        return None
    return value

def _cache_set(key: str, value: Any, ttl_seconds: int):
    _RESPONSE_CACHE[key] = (time.time() + ttl_seconds, value)
    return value

def _seq(name: str) -> int:
    doc = db.counters.find_one_and_update(
        {"_id": name}, {"$inc": {"sequence_value": 1}},
        upsert=True, return_document=True,
    )
    return doc["sequence_value"]

def next_ticket_number() -> str:
    year = datetime.now(timezone.utc).year
    seq = _seq(f"ticket_{year}")
    return f"TKT-{year}-{seq:04d}"

def next_device_id() -> str:
    year = datetime.now(timezone.utc).year
    seq = _seq(f"device_{year}")
    return f"DEV-{year}-{seq:04d}"

def next_company_code() -> str:
    max_retries = 100
    for attempt in range(max_retries):
        seq = _seq("company")
        code = f"CMP-{seq:04d}"
        if not db.companies.find_one({"company_code": code}):
            return code
    raise ValueError("Unable to generate unique company code")

# ---------- Models ----------
class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class OTPVerifyRequest(BaseModel):
    email: EmailStr
    otp: str
    challenge_id: str

class EngineerCreate(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None
    password: str
    employee_id: Optional[str] = None
    designation: Optional[str] = None
    address: Optional[str] = None
    is_remote: Optional[bool] = False
    oem_number: Optional[str] = None

class EngineerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    is_active: Optional[bool] = None
    is_available: Optional[bool] = None
    password: Optional[str] = None
    employee_id: Optional[str] = None
    designation: Optional[str] = None
    address: Optional[str] = None
    is_remote: Optional[bool] = None
    oem_number: Optional[str] = None

class CompanyCreate(BaseModel):
    company_name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[EmailStr] = None
    client_email: Optional[EmailStr] = None
    address: Optional[str] = None
    gst_number: Optional[str] = None
    product_ref_number: Optional[str] = None
    oem_ref_number: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None

class CompanyUpdate(BaseModel):
    company_name: Optional[str] = None
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[EmailStr] = None
    client_email: Optional[EmailStr] = None
    address: Optional[str] = None
    gst_number: Optional[str] = None
    product_ref_number: Optional[str] = None
    oem_ref_number: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    pincode: Optional[str] = None
    status: Optional[Literal["active", "inactive"]] = None

class DeviceCreate(BaseModel):
    brand: str
    model: str
    serial_number: Optional[str] = None
    device_name: Optional[str] = None
    device_type: Optional[str] = None
    warranty_status: Literal["active", "expired", "none"] = "none"
    warranty_expiry: Optional[str] = None
    purchase_date: Optional[str] = None
    notes: Optional[str] = None

class TicketCreate(BaseModel):
    company_id: str
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    customer_email: EmailStr
    contact_source: Literal["call", "whatsapp", "email"] = "call"
    issue_description: str
    priority: Literal["low", "medium", "high", "critical"] = "medium"
    product_reference_number: Optional[str] = None
    oem_reference_number: Optional[str] = None
    current_address: Optional[str] = None
    device: Optional[DeviceCreate] = None
    devices: Optional[List[DeviceCreate]] = None

class TicketAssign(BaseModel):
    engineer_id: Optional[str] = None
    is_outsource: bool = False
    outsource_name: Optional[str] = None
    outsource_company: Optional[str] = None
    outsource_phone: Optional[str] = None
    outsource_price: Optional[float] = None
    outsource_notes: Optional[str] = None

class ServiceReportCreate(BaseModel):
    work_done: str
    resolution_summary: Optional[str] = None
    parts_used: Optional[List[dict]] = []
    customer_signed_name: Optional[str] = None

class StatusUpdate(BaseModel):
    status: Literal[
        "accepted", "travelling", "reached_site", "in_progress",
        "resolved", "completed_with_signature", "report_generated",
        "closed", "rejected",
    ]
    note: Optional[str] = None
    reject_reason: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None

class LocationUpdate(BaseModel):
    latitude: float
    longitude: float

class PartItem(BaseModel):
    name: str
    part_number: Optional[str] = None
    quantity: int = 1

class ReportSubmit(BaseModel):
    engineer_notes: str
    resolution_summary: Optional[str] = None
    parts_used: List[PartItem] = []
    before_images: List[str] = []
    after_images: List[str] = []
    customer_signature: str
    customer_signed_name: Optional[str] = None

class AttendanceAction(BaseModel):
    latitude: Optional[float] = None
    longitude: Optional[float] = None

# ---------- Startup ----------
@app.on_event("startup")
async def startup():
    # ---- Migrations: backfill new fields on existing docs ----
    db.tickets.update_many(
        {"ticket_no": {"$exists": False}, "ticket_number": {"$exists": True}},
        [{"$set": {"ticket_no": "$ticket_number"}}],
    )
    db.tickets.update_many(
        {"issue_description": {"$exists": False}, "problem_description": {"$exists": True}},
        [{"$set": {"issue_description": "$problem_description"}}],
    )

    # Drop legacy indexes that conflict
    try:
        db.tickets.drop_index("ticket_number_1")
    except Exception:
        pass
    # Rebuild serial_number as a partial unique index. Older builds created a
    # plain unique index, which treats multiple null/missing values as duplicates.
    try:
        for idx in db.devices.list_indexes():
            if "serial_number" in idx.get("key", {}):
                db.devices.drop_index(idx["name"])
    except Exception:
        pass

    # Sync counters with existing data
    year = datetime.now(timezone.utc).year

    last_device = db.devices.find_one(
        {"device_id": {"$regex": f"^DEV-{year}-"}},
        sort=[("device_id", -1)]
    )

    if last_device:
        try:
            seq = int(last_device["device_id"].split("-")[-1])
            db.counters.update_one(
                {"_id": f"device_{year}"},
                {"$max": {"sequence_value": seq}},
                upsert=True,
            )
        except Exception:
            pass

    last_ticket = db.tickets.find_one(
        {"ticket_no": {"$regex": f"^TKT-{year}-"}},
        sort=[("ticket_no", -1)]
    )

    if last_ticket:
        try:
            seq = int(last_ticket["ticket_no"].split("-")[-1])
            db.counters.update_one(
                {"_id": f"ticket_{year}"},
                {"$max": {"sequence_value": seq}},
                upsert=True,
            )
        except Exception:
            pass

    # Create indexes only if they don't exist (background=True for non-blocking)
    _indexes_created = getattr(app.state, '_indexes_created', False)
    if not _indexes_created:
        # Users
        db.users.create_index("role")
        db.users.create_index("email", unique=True)
        db.users.create_index("id", unique=True)
        db.users.create_index("employee_id", sparse=True)

        # Companies
        db.companies.create_index("company_name", unique=True)
        db.companies.create_index("company_code", unique=True)
        db.companies.create_index("status")
        db.companies.create_index([("company_name", "text"), ("company_code", "text"), ("contact_person", "text"), ("city", "text")])

        # Tickets (most critical for performance)
        db.tickets.create_index("ticket_no", unique=True, sparse=True)
        db.tickets.create_index("status")
        db.tickets.create_index("assigned_engineer_id")
        db.tickets.create_index("company_id")
        db.tickets.create_index([("created_at", -1)])
        db.tickets.create_index([("company_id", 1), ("created_at", -1)])
        db.tickets.create_index([("status", 1), ("created_at", -1)])
        db.tickets.create_index([("is_deleted", 1), ("created_at", -1)])
        db.tickets.create_index([("assigned_engineer_id", 1), ("is_deleted", 1), ("created_at", -1)])
        db.tickets.create_index([("assigned_engineer_id", 1), ("status", 1), ("is_deleted", 1), ("created_at", -1)])
        db.tickets.create_index([("company_id", 1), ("status", 1), ("is_deleted", 1), ("created_at", -1)])
        db.tickets.create_index("is_deleted")

        # Devices
        db.devices.create_index("device_id", unique=True)
        db.devices.create_index(
            "serial_number",
            unique=True,
            partialFilterExpression={"serial_number": {"$type": "string"}},
        )
        db.devices.create_index("company_id")
        db.devices.create_index([("company_id", 1), ("created_at", -1)])
        db.devices.create_index([("warranty_status", 1), ("warranty_expiry", 1)])
        db.devices.create_index("brand")
        db.devices.create_index("model")
        db.devices.create_index("device_name")
        db.devices.create_index([("serial_number", "text"), ("device_id", "text"), ("brand", "text"), ("model", "text"), ("device_name", "text")])
        db.devices.create_index("warranty_expiry")

        # Other collections
        db.ticket_status_logs.create_index("ticket_id")
        db.ticket_status_logs.create_index([("timestamp", -1)])
        db.ticket_status_logs.create_index([("ticket_id", 1), ("timestamp", -1)])
        db.service_reports.create_index("ticket_id", unique=True)
        db.attachments.create_index("ticket_id")
        db.notifications.create_index("user_id")
        db.otp_challenges.create_index("expires_at", expireAfterSeconds=0)

        # Ticket issues
        db.ticket_issues.create_index("ticket_id")
        db.ticket_issues.create_index("engineer_id")
        db.ticket_issues.create_index("company_id")
        db.ticket_issues.create_index([("status", 1), ("created_at", -1)])
        db.ticket_issues.create_index([("created_at", -1)])

        app.state._indexes_created = True
        logger.info("✓ Database indexes created")

    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "vishal@plutusventures.in")
    admin_password = os.environ.get("ADMIN_PASSWORD", "Vishal@PV2026")

    existing = db.users.find_one({"email": admin_email})

    if not existing:
        db.users.insert_one({
            "id": new_id(),
            "email": admin_email,
            "name": "Administrator",
            "role": "admin",
            "password_hash": hash_password(admin_password),
            "is_active": True,
            "is_available": True,
            "status": "active",
            "designation": "System Admin",
            "created_at": now_iso(),
            "updated_at": now_iso(),
        })
        logger.info(f"Seeded admin: {admin_email}")
    else:
        if not verify_password(admin_password, existing["password_hash"]):
            db.users.update_one(
                {"email": admin_email},
                {
                    "$set": {
                        "password_hash": hash_password(admin_password),
                        "updated_at": now_iso()
                    }
                }
            )

    # Also keep legacy admin@serviceops.com if it exists
    # legacy = db.users.find_one({"email": "admin@serviceops.com"})
    # if not legacy:
    #     db.users.insert_one({
    #         "id": new_id(),
    #         "email": "admin@serviceops.com",
    #         "name": "Admin (legacy)",
    #         "role": "admin",
    #         "password_hash": hash_password("admin123"),
    #         "is_active": True,
    #         "is_available": True,
    #         "status": "active",
    #         "created_at": now_iso(),
    #         "updated_at": now_iso(),
    #     })

    # Seed engineer
    # eng_email = "engineer@plutusventures.com"
    # if not db.users.find_one({"email": eng_email}):
    #     db.users.insert_one({
    #         "id": new_id(),
    #         "email": eng_email,
    #         "name": "Rajiv Kumar",
    #         "role": "engineer",
    #         "phone": "+91 98765 43210",
    #         "password_hash": hash_password("engineer123"),
    #         "skills": ["Laptop Repair", "Networking", "Printer"],
    #         "employee_id": "EMP-001",
    #         "designation": "Senior Field Engineer",
    #         "is_active": True,
    #         "is_available": True,
    #         "status": "active",
    #         "created_at": now_iso(),
    #         "updated_at": now_iso(),
    #     })

    # Legacy engineer
    # legacy_eng = db.users.find_one({"email": "engineer@serviceops.com"})
    # if not legacy_eng:
    #     db.users.insert_one({
    #         "id": new_id(),
    #         "email": "engineer@serviceops.com",
    #         "name": "Field Engineer (legacy)",
    #         "role": "engineer",
    #         "phone": "+91 90000 00000",
    #         "password_hash": hash_password("engineer123"),
    #         "skills": ["Laptop Repair"],
    #         "employee_id": "EMP-LEG",
    #         "is_active": True,
    #         "is_available": True,
    #         "status": "active",
    #         "created_at": now_iso(),
    #         "updated_at": now_iso(),
    #     })

    # Init local filesystem storage
    init_storage()

    # Write test credentials
    from pathlib import Path

    try:
        BASE_DIR = Path(__file__).resolve().parent
        memory_dir = BASE_DIR / "memory"
        memory_dir.mkdir(parents=True, exist_ok=True)

        (memory_dir / "test_credentials.md").write_text(
            "# Test Credentials\n\n"
            "## Admin (primary)\n"
            f"- Email: `{admin_email}`\n"
            f"- Password: `{admin_password}`\n\n"
            "## Admin (legacy)\n"
            "- Email: `admin@serviceops.com` / Password: `admin123`\n\n"
            "## Engineer (primary)\n"
            f"- Email: `{eng_email}` / Password: `engineer123`\n\n"
            "## Engineer (legacy)\n"
            "- Email: `engineer@serviceops.com` / Password: `engineer123`\n\n"
            "## Login flow\n"
            "1. POST /api/auth/login -> returns `challenge_id` and `dev_otp`\n"
            "2. POST /api/auth/verify-otp -> returns JWT `token`\n"
            "3. Use `Authorization: Bearer <token>` for all subsequent calls\n"
        )

    except Exception as e:
        logger.error(f"Error writing test credentials: {e}")
     
@app.on_event("shutdown")
async def shutdown():
    client.close()

# ---------- AUTH ----------
@api.post("/auth/login")
@limiter.limit("10/minute")
async def login(request: Request, payload: LoginRequest):
    email = payload.email.lower().strip()
    user = db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="Account disabled")

    # Update last login
    db.users.update_one({"email": email}, {"$set": {"last_login": now_iso()}})

    token = create_access_token(
        user.get("id") or str(user["_id"]),
        user["email"],
        user.get("role", "admin"),
    )
    return {
        "token": token,
        "user": {
            "id": user.get("id") or str(user["_id"]),
            "name": user.get("name", ""),
            "email": user["email"],
            "role": user.get("role", "admin"),
        },
    }
@api.post("/auth/verify-otp")
@limiter.limit("20/minute")
async def verify_otp(request: Request, payload: OTPVerifyRequest):
    try:
        email = payload.email.lower().strip()

        #  Fetch challenge
        challenge = db.otp_challenges.find_one({
            "id": payload.challenge_id,
            "email": email,
            "consumed": False,
        })

        if not challenge:
            raise HTTPException(status_code=400, detail="Invalid challenge")

        #  OTP match
        if str(challenge.get("otp")) != payload.otp.strip():
            raise HTTPException(status_code=400, detail="Incorrect OTP")
        
        #EXPIRY CHECK (SAFE)
        expires_at = challenge.get("expires_at")
        from datetime import timezone
        if expires_at:
            if expires_at.replace(tzinfo=timezone.utc) < datetime.now(timezone.utc):
                raise HTTPException(status_code=400, detail="OTP expired")
  
        #  Mark consumed
        db.otp_challenges.update_one(
            {"id": payload.challenge_id},
            {"$set": {"consumed": True}}
        )

        # Fetch full user
        user = db.users.find_one({"email": email})
        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        #  Update last login
        db.users.update_one(
            {"email": email},
            {"$set": {"last_login": now_iso()}}
        )

        #  Create token (SAFE format)
        token = create_access_token(
            str(user["_id"]),
            user["email"],
            user.get("role", "admin")
        )

        return {
            "token": token,
            "user": {
                "email": user["email"],
                "role": user.get("role", "admin")
            }
        }

    except Exception as e:
        print("VERIFY OTP ERROR:", str(e))  # 👈 VERY IMPORTANT
        raise HTTPException(status_code=500, detail="Internal error")

@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user

@api.post("/auth/logout")
async def logout(user=Depends(get_current_user)):
    return {"ok": True}

class FCMTokenUpdate(BaseModel):
    token: str

@api.post("/users/fcm-token")
async def save_fcm_token(payload: FCMTokenUpdate, user=Depends(get_current_user)):
    db.users.update_one(
        {"id": user["id"]},
        {"$set": {"fcm_token": payload.token, "updated_at": now_iso()}}
    )
    return {"status": "saved"}

# ---------- ENGINEERS ----------
@api.get("/engineers")
async def list_engineers(available_only: bool = False,
                         user=Depends(get_current_user)):
    q = {"role": "engineer"}
    if available_only:
        q["is_active"] = True
        q["is_available"] = True
    engs = list(db.users.find(q, {"_id": 0, "password_hash": 0}))
    eng_ids = [e["id"] for e in engs]
    active_tickets = db.tickets.aggregate([
        {"$match": {"assigned_engineer_id": {"$in": eng_ids}, "status": {"$nin": ["closed", "rejected", "report_generated"]}}},
        {"$group": {"_id": "$assigned_engineer_id", "count": {"$sum": 1}}}
    ])
    ticket_counts = {doc["_id"]: doc["count"] for doc in active_tickets}
    for e in engs:
        e["active_tickets"] = ticket_counts.get(e["id"], 0)
    return engs

@api.post("/engineers", dependencies=[Depends(require_admin)])
async def create_engineer(payload: EngineerCreate):
    email = payload.email.lower().strip()
    if db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")
    doc = {
        "id": new_id(),
        "email": email,
        "name": payload.name,
        "role": "engineer",
        "phone": payload.phone,
        "skills": [],
        "employee_id": payload.employee_id,
        "designation": payload.designation,
        "address": payload.address,
        "is_remote": payload.is_remote or False,
        "oem_number": payload.oem_number,
        "password_hash": hash_password(payload.password),
        "is_active": True,
        "is_available": True,
        "status": "active",
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    db.users.insert_one(doc)
    return clean({**doc})

@api.patch("/engineers/{eng_id}")
async def update_engineer(eng_id: str, payload: EngineerUpdate,
                          current_user=Depends(get_current_user)):
    role = current_user.get("role")
    # Engineers can only update their own is_remote and oem_number
    if role == "engineer":
        if current_user.get("id") != eng_id:
            raise HTTPException(status_code=403, detail="Cannot update another engineer")
        allowed = {"is_remote", "oem_number"}
        updates = {k: v for k, v in payload.model_dump().items()
                   if k in allowed and v is not None}
        # Allow setting is_remote to False (it's falsy so needs special handling)
        if payload.is_remote is not None:
            updates["is_remote"] = payload.is_remote
    elif role in ("admin", "sub_admin"):
        updates = {k: v for k, v in payload.model_dump().items() if v is not None}
        if "password" in updates:
            updates["password_hash"] = hash_password(updates.pop("password"))
    else:
        raise HTTPException(status_code=403, detail="Access denied")
    updates["updated_at"] = now_iso()
    if len(updates) <= 1:
        raise HTTPException(status_code=400, detail="No changes")
    res = db.users.update_one({"id": eng_id, "role": "engineer"},
                               {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Engineer not found")
    return db.users.find_one({"id": eng_id}, {"_id": 0, "password_hash": 0})

@api.delete("/engineers/{eng_id}", dependencies=[Depends(require_admin)])
async def delete_engineer(eng_id: str):
    res = db.users.delete_one({"id": eng_id, "role": "engineer"})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Engineer not found")
    return {"ok": True}

# ---------- ADMINS ----------
class AdminCreate(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None
    password: str
    designation: Optional[str] = None

class TicketAdminCreate(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None
    password: str
    designation: Optional[str] = None

class TicketAdminUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    password: Optional[str] = None
    designation: Optional[str] = None
    is_active: Optional[bool] = None

@api.get("/admins", dependencies=[Depends(require_admin)])
async def list_admins():
    items = list(db.users.find({"role": "admin"}, {"_id": 0, "password_hash": 0}).sort("created_at", -1))
    return {"items": items, "total": len(items)}

@api.post("/admins", dependencies=[Depends(require_admin)])
async def create_admin(payload: AdminCreate):
    email = payload.email.lower().strip()
    if db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")
    doc = {
        "id": new_id(),
        "name": payload.name,
        "email": email,
        "phone": payload.phone,
        "designation": payload.designation or "Admin",
        "role": "admin",
        "password_hash": hash_password(payload.password),
        "is_active": True,
        "is_available": True,
        "status": "active",
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    db.users.insert_one(doc)
    return clean({**doc})

@api.delete("/admins/{admin_id}", dependencies=[Depends(require_admin)])
async def delete_admin(admin_id: str):
    res = db.users.delete_one({"id": admin_id, "role": "admin"})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Admin not found")
    return {"status": "deleted"}

# ---------- TICKET ADMINS ----------
@api.get("/ticket-admins", dependencies=[Depends(require_admin)])
async def list_ticket_admins(q: Optional[str] = None):
    query: Dict[str, Any] = {"role": "ticket_admin"}
    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"email": {"$regex": q, "$options": "i"}},
        ]
    items = list(db.users.find(query, {"_id": 0, "password_hash": 0}).sort("created_at", -1))
    return {"items": items, "total": len(items)}

@api.post("/ticket-admins", dependencies=[Depends(require_admin)])
async def create_ticket_admin(payload: TicketAdminCreate):
    email = payload.email.lower().strip()
    if db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")
    doc = {
        "id": new_id(),
        "name": payload.name,
        "email": email,
        "phone": payload.phone,
        "designation": payload.designation or "Ticket Admin",
        "role": "ticket_admin",
        "password_hash": hash_password(payload.password),
        "is_active": True,
        "is_available": True,
        "status": "active",
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    db.users.insert_one(doc)
    return clean({**doc})

@api.patch("/ticket-admins/{ticket_admin_id}", dependencies=[Depends(require_admin)])
async def update_ticket_admin(ticket_admin_id: str, payload: TicketAdminUpdate):
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if "password" in updates:
        updates["password_hash"] = hash_password(updates.pop("password"))
    updates["updated_at"] = now_iso()
    if len(updates) == 1:
        raise HTTPException(status_code=400, detail="No changes provided")
    res = db.users.update_one(
        {"id": ticket_admin_id, "role": "ticket_admin"},
        {"$set": updates},
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ticket admin not found")
    return clean(db.users.find_one(
        {"id": ticket_admin_id},
        {"_id": 0, "password_hash": 0},
    ))

@api.delete("/ticket-admins/{ticket_admin_id}", dependencies=[Depends(require_admin)])
async def delete_ticket_admin(ticket_admin_id: str):
    res = db.users.delete_one({"id": ticket_admin_id, "role": "ticket_admin"})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ticket admin not found")
    return {"ok": True}

# ---------- SUB-ADMINS ----------
class SubAdminCreate(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None
    password: str
    employee_id: Optional[str] = None
    designation: Optional[str] = None
    address: Optional[str] = None

class SubAdminUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None
    employee_id: Optional[str] = None
    designation: Optional[str] = None
    address: Optional[str] = None
    # NOTE: assigned_company_ids is intentionally excluded here.
    # Only the sub-admin themselves can manage their own company assignments
    # via PATCH /sub-admins/me/companies

class SubAdminSelfCompaniesUpdate(BaseModel):
    assigned_company_ids: List[str] = []

@api.get("/sub-admins", dependencies=[Depends(require_admin)])
async def list_sub_admins(q: Optional[str] = None):
    query: Dict[str, Any] = {"role": "sub_admin"}
    if q:
        query["$or"] = [
            {"name": {"$regex": q, "$options": "i"}},
            {"email": {"$regex": q, "$options": "i"}},
            {"employee_id": {"$regex": q, "$options": "i"}},
        ]
    items = list(db.users.find(query, {"_id": 0, "password_hash": 0}).sort("created_at", -1))
    return {"items": items, "total": len(items)}

@api.post("/sub-admins", dependencies=[Depends(require_admin)])
async def create_sub_admin(payload: SubAdminCreate):
    email = payload.email.lower().strip()
    if db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")
    doc = {
        "id": new_id(),
        "email": email,
        "name": payload.name,
        "role": "sub_admin",
        "phone": payload.phone,
        "employee_id": payload.employee_id,
        "designation": payload.designation,
        "address": payload.address,
        "password_hash": hash_password(payload.password),
        "is_active": True,
        "status": "active",
        "assigned_company_ids": [],
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    db.users.insert_one(doc)
    return clean({**doc})

@api.patch("/sub-admins/{sub_admin_id}", dependencies=[Depends(require_admin)])
async def update_sub_admin(sub_admin_id: str, payload: SubAdminUpdate):
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if "password" in updates:
        updates["password_hash"] = hash_password(updates.pop("password"))
    updates["updated_at"] = now_iso()
    if len(updates) == 1:  # only updated_at
        raise HTTPException(status_code=400, detail="No changes provided")
    res = db.users.update_one({"id": sub_admin_id, "role": "sub_admin"}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sub-admin not found")
    return clean(db.users.find_one({"id": sub_admin_id}, {"_id": 0, "password_hash": 0}))

@api.patch("/sub-admins/me/companies")
async def update_my_companies(
    payload: SubAdminSelfCompaniesUpdate,
    user=Depends(get_current_user),
):
    """Sub-admin self-service: update their own assigned company list.
    Admin role is explicitly blocked — only sub_admin may call this."""
    if user.get("role") != "sub_admin":
        raise HTTPException(
            status_code=403,
            detail="Only sub-admins can manage their own company assignments",
        )
    updated_at = now_iso()
    db.users.update_one(
        {"id": user["id"], "role": "sub_admin"},
        {"$set": {
            "assigned_company_ids": payload.assigned_company_ids,
            "updated_at": updated_at,
        }},
    )
    db.ticket_status_logs.insert_one({
        "id": new_id(),
        "ticket_id": None,
        "actor_type": "sub_admin_assignment",
        "target_user_id": user["id"],
        "target_user_name": user.get("name"),
        "old_companies": user.get("assigned_company_ids", []),
        "new_companies": payload.assigned_company_ids,
        "changed_by": user["id"],
        "changed_by_name": user.get("name"),
        "changed_by_role": user.get("role"),
        "timestamp": updated_at,
    })
    return clean(db.users.find_one({"id": user["id"]}, {"_id": 0, "password_hash": 0}))

@api.delete("/sub-admins/{sub_admin_id}", dependencies=[Depends(require_admin)])
async def delete_sub_admin(sub_admin_id: str):
    res = db.users.delete_one({"id": sub_admin_id, "role": "sub_admin"})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sub-admin not found")
    return {"ok": True}

# ---------- COMPANIES ----------
@api.get("/companies")
async def list_companies(
    q: Optional[str] = None,
    status: Optional[Literal["active", "inactive"]] = None,
    page: int = 1, page_size: int = 50,
    user=Depends(get_current_user),
):
    cache_k = cache_key("companies", status=status or "", q=q or "", page=page, page_size=page_size)
    cached = get_cache(cache_k, max_age_seconds=300)
    if cached:
        return cached

    query: Dict[str, Any] = {}
    if status:
        query["status"] = status
    if q:
        query["$text"] = {"$search": q}

    skip = max(0, (page - 1)) * page_size

    # Single aggregation instead of 2 separate queries
    pipeline = [
        {"$match": query},
        {
            "$facet": {
                "count": [{"$count": "total"}],
                "items": [
                    {"$sort": {"created_at": -1}},
                    {"$skip": skip},
                    {"$limit": page_size},
                    {"$project": {"_id": 0}},
                ]
            }
        }
    ]
    result = list(db.companies.aggregate(pipeline))
    if result:
        r = result[0]
        total = r["count"][0]["total"] if r["count"] else 0
        items = r["items"]
    else:
        total = 0
        items = []

    response = {"items": items, "total": total, "page": page, "page_size": page_size}
    set_cache(cache_k, response)
    return response

@api.get("/companies/{company_id}")
async def get_company(company_id: str, user=Depends(get_current_user)):
    c = db.companies.find_one({"id": company_id}, {"_id": 0})
    if not c:
        raise HTTPException(status_code=404, detail="Company not found")
    tickets = list(db.tickets.find(
        {"company_id": company_id}, {"_id": 0}
    ).sort("created_at", -1).limit(20))
    devices = list(db.devices.find(
        {"company_id": company_id}, {"_id": 0}
    ).sort("created_at", -1).limit(50))
    return {"company": c, "tickets": tickets, "devices": devices}

@api.post("/companies", dependencies=[Depends(require_sub_admin)])
async def create_company(payload: CompanyCreate, admin=Depends(require_sub_admin)):
    name = payload.company_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Company name required")
    if db.companies.find_one({"company_name": {"$regex": f"^{name}$", "$options": "i"}}):
        raise HTTPException(status_code=400, detail="Company name already exists")
    doc = {
        "id": new_id(),
        "company_name": name,
        "company_code": next_company_code(),
        "contact_person": payload.contact_person,
        "phone": payload.phone,
        "email": payload.email,
        "address": payload.address,
        "gst_number": payload.gst_number,
        "product_ref_number": payload.product_ref_number,
        "oem_ref_number": payload.oem_ref_number,
        "client_email": payload.client_email,
        "city": payload.city,
        "state": payload.state,
        "pincode": payload.pincode,
        "status": "active",
        "created_by": admin["id"],
        "created_by_name": admin.get("name", ""),
        "created_by_role": admin.get("role", ""),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    db.companies.insert_one(doc)
    return clean({**doc})

@api.post("/companies/bulk-import", dependencies=[Depends(require_sub_admin)])
async def bulk_import_companies(
    file: UploadFile = File(...),
    admin=Depends(require_sub_admin),
):
    """Import companies from an Excel (.xlsx) or CSV file."""
    content = await file.read()
    filename = (file.filename or "").lower()

    rows = []
    if filename.endswith(".csv"):
        import csv, codecs
        reader = csv.DictReader(codecs.iterdecode(io.BytesIO(content), "utf-8-sig"))
        rows = list(reader)
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
    else:
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv files are supported")

    # Column name aliases
    ALIASES = {
        "name": "company_name", "company": "company_name",
        "contact": "contact_person", "gst": "gst_number",
        "pin": "pincode", "zip": "pincode",
    }

    created, skipped, errors = [], [], []
    for i, raw in enumerate(rows, start=2):
        row = {ALIASES.get(k, k): v for k, v in raw.items()}
        name = (row.get("company_name") or "").strip()
        if not name:
            continue
        if db.companies.find_one({"company_name": {"$regex": f"^{name}$", "$options": "i"}}):
            skipped.append(name)
            continue
        try:
            doc = {
                "id": new_id(),
                "company_name": name,
                "company_code": next_company_code(),
                "contact_person": row.get("contact_person") or None,
                "phone": row.get("phone") or None,
                "email": row.get("email") or None,
                "address": row.get("address") or None,
                "gst_number": row.get("gst_number") or None,
                "city": row.get("city") or None,
                "state": row.get("state") or None,
                "pincode": row.get("pincode") or None,
                "status": "active",
                "created_by": admin["id"],
                "created_by_name": admin.get("name", ""),
                "created_by_role": admin.get("role", ""),
                "created_at": now_iso(),
                "updated_at": now_iso(),
            }
            db.companies.insert_one(doc)
            created.append(name)
        except Exception as e:
            errors.append({"row": i, "name": name, "error": str(e)})

    return {"created": len(created), "skipped": len(skipped), "errors": errors,
            "skipped_names": skipped}

@api.put("/companies/{company_id}", dependencies=[Depends(require_sub_admin)])
async def update_company(company_id: str, payload: CompanyUpdate):
    existing = db.companies.find_one({"id": company_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Company not found")
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if "company_name" in updates:
        new_name = updates["company_name"].strip()
        if not new_name:
            raise HTTPException(status_code=400, detail="Company name required")
        dup = db.companies.find_one({
            "company_name": {"$regex": f"^{new_name}$", "$options": "i"},
            "id": {"$ne": company_id},
        })
        if dup:
            raise HTTPException(status_code=400, detail="Company name already exists")
        updates["company_name"] = new_name
    updates["updated_at"] = now_iso()
    db.companies.update_one({"id": company_id}, {"$set": updates})
    if "company_name" in updates:
        db.tickets.update_many({"company_id": company_id},
                               {"$set": {"company_name": updates["company_name"]}})
        db.devices.update_many({"company_id": company_id},
                               {"$set": {"company_name": updates["company_name"]}})
    return db.companies.find_one({"id": company_id}, {"_id": 0})

@api.delete("/companies/{company_id}", dependencies=[Depends(require_admin)])
async def delete_company(company_id: str):
    open_tix = db.tickets.count_documents({
        "company_id": company_id,
        "status": {"$nin": ["closed", "rejected"]},
    })
    if open_tix > 0:
        raise HTTPException(status_code=400,
                            detail=f"Cannot delete: {open_tix} active tickets")
    res = db.companies.delete_one({"id": company_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    return {"ok": True}

# ---------- DEVICES ----------
@api.get("/devices")
async def list_devices(q: Optional[str] = None,
                       company_id: Optional[str] = None,
                       page: int = Query(1, ge=1),
                       per_page: int = Query(100, ge=1, le=500),
                       user=Depends(get_current_user)):
    query: Dict[str, Any] = {}
    if company_id:
        query["company_id"] = company_id
    if q:
        query["$text"] = {"$search": q}

    skip = (page - 1) * per_page

    # Single aggregation instead of 2 separate queries
    pipeline = [
        {"$match": query},
        {
            "$facet": {
                "count": [{"$count": "total"}],
                "items": [
                    {"$sort": {"created_at": -1}},
                    {"$skip": skip},
                    {"$limit": per_page},
                    {"$project": {
                        "_id": 0,
                        "id": 1,
                        "device_id": 1,
                        "company_id": 1,
                        "company_name": 1,
                        "brand": 1,
                        "model": 1,
                        "device_name": 1,
                        "device_type": 1,
                        "serial_number": 1,
                        "warranty_status": 1,
                        "warranty_expiry": 1,
                        "purchase_date": 1,
                        "created_at": 1,
                    }},
                ]
            }
        }
    ]
    result = list(db.devices.aggregate(pipeline))
    if result:
        r = result[0]
        total = r["count"][0]["total"] if r["count"] else 0
        items = r["items"]
    else:
        total = 0
        items = []

    return {"items": items, "total": total, "page": page, "per_page": per_page}

@api.get("/devices/history-export")
async def export_device_history(
    company_id: Optional[str] = None,
    month_from: Optional[str] = None,
    month_to: Optional[str] = None,
    user=Depends(get_current_user)
):
    """Export device service history as Excel."""
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    ticket_filter = {}
    if company_id:
        ticket_filter["company_id"] = company_id
    if month_from:
        ticket_filter["created_at"] = {"$gte": month_from + "-01"}
    if month_to:
        import calendar
        y, m = int(month_to[:4]), int(month_to[5:7])
        last_day = calendar.monthrange(y, m)[1]
        ticket_filter.setdefault("created_at", {})["$lte"] = f"{month_to}-{last_day}"

    tickets = list(db.tickets.find(ticket_filter, {"_id": 0}).sort("created_at", -1).limit(1000))
    device_ids = list({
        dev_id
        for t in tickets
        for dev_id in _ticket_device_ids(t)
    })
    company_ids = list({t.get("company_id") for t in tickets if t.get("company_id")})
    eng_ids = list({t.get("assigned_engineer_id") for t in tickets if t.get("assigned_engineer_id")})

    devices_map = {d["device_id"]: d for d in db.devices.find({"device_id": {"$in": device_ids}}, {"_id": 0})} if device_ids else {}
    companies_map = {c["id"]: c for c in db.companies.find({"id": {"$in": company_ids}}, {"_id": 0, "id": 1, "company_name": 1})} if company_ids else {}
    engineers_map = {e["id"]: e["name"] for e in db.users.find({"id": {"$in": eng_ids}, "role": "engineer"}, {"_id": 0, "id": 1, "name": 1})} if eng_ids else {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Device History"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0A1128")
    headers = ["Ticket No", "Device ID", "Brand", "Model", "Serial No",
               "Company", "Engineer", "Status", "Issue", "Created At", "Completed At"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, t in enumerate(tickets, 2):
        ticket_devices = [
            devices_map[dev_id]
            for dev_id in _ticket_device_ids(t)
            if dev_id in devices_map
        ]
        ws.cell(row=row, column=1, value=t.get("ticket_no") or t.get("ticket_number"))
        ws.cell(row=row, column=2, value=", ".join([d.get("device_id", "") for d in ticket_devices]) or t.get("device_id"))
        ws.cell(row=row, column=3, value=", ".join([d.get("brand", "") for d in ticket_devices]))
        ws.cell(row=row, column=4, value=", ".join([d.get("model", "") for d in ticket_devices]))
        ws.cell(row=row, column=5, value=", ".join([d.get("serial_number", "") for d in ticket_devices if d.get("serial_number")]))
        ws.cell(row=row, column=6, value=companies_map.get(t.get("company_id"), {}).get("company_name"))
        ws.cell(row=row, column=7, value=engineers_map.get(t.get("assigned_engineer_id")))
        ws.cell(row=row, column=8, value=t.get("status"))
        ws.cell(row=row, column=9, value=t.get("issue_description") or t.get("problem_description"))
        ws.cell(row=row, column=10, value=(t.get("created_at") or "")[:16].replace("T", " "))
        ws.cell(row=row, column=11, value=(t.get("completed_at") or "")[:16].replace("T", " "))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=device_history.xlsx"}
    )

# ---------- DEVICE HISTORY (Feature 3) ----------
def _build_device_history_query(company: Optional[str], start_date: Optional[str],
                                end_date: Optional[str],
                                include_deleted: bool = False,
                                only_deleted: bool = False) -> Dict[str, Any]:
    """Build the MongoDB query for device-history listings.

    company: company name (case-insensitive partial match)
    start_date / end_date: YYYY-MM-DD inclusive (matched against ticket.created_at ISO string).
    include_deleted: if True, return both active and soft-deleted records.
    only_deleted: if True, return only soft-deleted records.
    Default (include_deleted=False, only_deleted=False) excludes soft-deleted records.
    """
    q: Dict[str, Any] = {}
    if only_deleted:
        q["is_deleted"] = True
    elif not include_deleted:
        q["is_deleted"] = {"$ne": True}

    q["status"] = {"$in": ["assigned", "closed", "rejected", "report_generated", "completed_with_signature"]}

    if company:
        company_doc_ids = [
            c["id"] for c in db.companies.find(
                {"company_name": {"$regex": company, "$options": "i"}},
                {"_id": 0, "id": 1},
            )
        ]
        # Match either company_id (preferred) or denormalised company_name on ticket.
        q["$or"] = [
            {"company_id": {"$in": company_doc_ids}} if company_doc_ids else {"_no_match_": True},
            {"company_name": {"$regex": company, "$options": "i"}},
        ]
    if start_date:
        q.setdefault("created_at", {})["$gte"] = start_date
    if end_date:
        # Inclusive upper bound up to end of day.
        q.setdefault("created_at", {})["$lte"] = f"{end_date}T23:59:59"
    return q

def _enrich_history_rows(tickets: List[dict]) -> List[dict]:
    """Format ticket rows without N+1 queries - assume data already denormalized."""
    rows = []
    for t in tickets:
        rows.append({
            "id": t.get("id"),
            "device_id": ", ".join(_ticket_device_ids(t)) or t.get("device_id"),
            "ticket_id": t.get("ticket_no") or t.get("ticket_number") or t.get("id"),
            "company_name": t.get("company_name", "—"),
            "engineer_name": t.get("assigned_engineer_name", "—"),
            "status": t.get("status"),
            "created_date": (t.get("created_at") or "")[:19].replace("T", " "),
            "closed_date": (
                (t.get("approved_at") or t.get("completed_at") or "")[:19].replace("T", " ")
                if t.get("status") == "closed" else ""
            ),
            "product_reference_number": t.get("product_reference_number", "—"),
            "oem_reference_number": t.get("oem_reference_number", "—"),
            "is_deleted": bool(t.get("is_deleted", False)),
        })
    return rows

@api.get("/device-history")
async def list_device_history(
    company: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    include_deleted: bool = False,
    only_deleted: bool = False,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    user=Depends(require_ticket_operator),
):
    """List device service history with pagination."""
    if user.get("role") == "ticket_admin":
        include_deleted = False
        only_deleted = False
    q = _build_device_history_query(company, start_date, end_date,
                                    include_deleted=include_deleted,
                                    only_deleted=only_deleted)
    if user.get("role") == "sub_admin":
        assigned_ids = user.get("assigned_company_ids") or []
        scope = [{"created_by": user["id"]}]
        if assigned_ids:
            scope.append({"company_id": {"$in": assigned_ids}})
        q["$or"] = scope

    skip = (page - 1) * per_page

    # Single aggregation instead of 2 separate queries
    pipeline = [
        {"$match": q},
        {
            "$facet": {
                "count": [{"$count": "total"}],
                "items": [
                    {"$sort": {"created_at": -1}},
                    {"$skip": skip},
                    {"$limit": per_page},
                    {"$project": {
                        "_id": 0,
                        "id": 1,
                        "ticket_no": 1,
                        "ticket_number": 1,
                        "company_id": 1,
                        "company_name": 1,
                        "customer_name": 1,
                        "customer_phone": 1,
                        "customer_email": 1,
                        "customer_company": 1,
                        "priority": 1,
                        "device_id": 1,
                        "device_ids": 1,
                        "status": 1,
                        "assigned_engineer_id": 1,
                        "assigned_engineer_name": 1,
                        "created_by": 1,
                        "created_at": 1,
                        "updated_at": 1,
                    }},
                ]
            }
        }
    ]
    result = list(db.tickets.aggregate(pipeline))
    if result:
        r = result[0]
        total = r["count"][0]["total"] if r["count"] else 0
        items = r["items"]
    else:
        total = 0
        items = []

    return {"items": _enrich_history_rows(items), "total": total, "page": page, "per_page": per_page}

@api.get("/device-history/filter")
async def filter_device_history(
    company: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    include_deleted: bool = False,
    only_deleted: bool = False,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    user=Depends(require_ticket_operator),
):
    """Filter device history by company and date range with pagination."""
    if user.get("role") == "ticket_admin":
        include_deleted = False
        only_deleted = False
    q = _build_device_history_query(company, start_date, end_date,
                                    include_deleted=include_deleted,
                                    only_deleted=only_deleted)
    if user.get("role") == "sub_admin":
        assigned_ids = user.get("assigned_company_ids") or []
        scope = [{"created_by": user["id"]}]
        if assigned_ids:
            scope.append({"company_id": {"$in": assigned_ids}})
        q["$or"] = scope

    skip = (page - 1) * per_page

    # Single aggregation instead of 2 separate queries
    pipeline = [
        {"$match": q},
        {
            "$facet": {
                "count": [{"$count": "total"}],
                "items": [
                    {"$sort": {"created_at": -1}},
                    {"$skip": skip},
                    {"$limit": per_page},
                    {"$project": {
                        "_id": 0,
                        "id": 1,
                        "device_id": 1,
                        "device_ids": 1,
                        "device_count": 1,
                        "ticket_no": 1,
                        "ticket_number": 1,
                        "customer_name": 1,
                        "customer_phone": 1,
                        "customer_company": 1,
                        "created_by": 1,
                        "company_name": 1,
                        "assigned_engineer_name": 1,
                        "assigned_engineer_id": 1,
                        "status": 1,
                        "created_at": 1,
                        "approved_at": 1,
                        "completed_at": 1,
                        "product_reference_number": 1,
                        "oem_reference_number": 1,
                        "is_deleted": 1,
                    }},
                ]
            }
        }
    ]
    result = list(db.tickets.aggregate(pipeline))
    if result:
        r = result[0]
        total = r["count"][0]["total"] if r["count"] else 0
        items = r["items"]
    else:
        total = 0
        items = []

    return {"items": _enrich_history_rows(items), "total": total, "page": page, "per_page": per_page}

@api.get("/device-history/export")
async def export_device_history_v2(
    company: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user=Depends(require_ticket_operator),
):
    """Export filtered device history to Excel (.xlsx).

    Columns: Device ID, Ticket ID, Company Name, Engineer Name, Status,
    Created Date, Closed Date, Product Reference Number, OEM Reference Number.
    """
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    q = _build_device_history_query(company, start_date, end_date)
    if user.get("role") == "sub_admin":
        assigned_ids = user.get("assigned_company_ids") or []
        scope = [{"created_by": user["id"]}]
        if assigned_ids:
            scope.append({"company_id": {"$in": assigned_ids}})
        q["$or"] = scope
    tickets = list(db.tickets.find(q, {"_id": 0}).sort("created_at", -1).limit(5000))
    rows = _enrich_history_rows(tickets)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Device History"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0A1128")
    headers = [
        "Device ID", "Ticket ID", "Company Name", "Engineer Name", "Status",
        "Created Date", "Closed Date",
        "Product Reference Number", "OEM Reference Number",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for ridx, r in enumerate(rows, 2):
        ws.cell(row=ridx, column=1, value=r.get("device_id"))
        ws.cell(row=ridx, column=2, value=r.get("ticket_id"))
        ws.cell(row=ridx, column=3, value=r.get("company_name"))
        ws.cell(row=ridx, column=4, value=r.get("engineer_name"))
        ws.cell(row=ridx, column=5, value=r.get("status"))
        ws.cell(row=ridx, column=6, value=r.get("created_date"))
        ws.cell(row=ridx, column=7, value=r.get("closed_date"))
        ws.cell(row=ridx, column=8, value=r.get("product_reference_number"))
        ws.cell(row=ridx, column=9, value=r.get("oem_reference_number"))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    def _slug(s: Optional[str]) -> str:
        if not s:
            return "all"
        # Preserve ISO date format (YYYY-MM-DD) for date params.
        import re
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", s):
            return s
        return "".join(ch if ch.isalnum() else "_" for ch in s).strip("_") or "all"

    filename = (
        f"device_history_{_slug(company)}_"
        f"{_slug(start_date)}_{_slug(end_date)}.xlsx"
    )

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@api.delete("/device-history/{ticket_id}")
async def soft_delete_device_history(ticket_id: str, user=Depends(require_sub_admin)):
    """Soft-delete a device-history entry (ticket).

    Sets is_deleted=true so the record is hidden from device-history listings
    and exports. Active (open) tickets cannot be deleted to avoid impacting
    ongoing work; only completed / closed / rejected tickets are eligible.
    """
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        # Allow lookup by ticket_no as fallback
        ticket = db.tickets.find_one({"ticket_no": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Device history record not found")

    if user.get("role") == "sub_admin":
        company_ids = user.get("assigned_company_ids") or []
        if ticket.get("company_id") not in company_ids:
            raise HTTPException(status_code=403, detail="Not authorized to delete this record")

    if ticket.get("status") not in (
        "closed", "rejected", "report_generated", "completed_with_signature"
    ):
        raise HTTPException(
            status_code=400,
            detail="Only completed or closed tickets can be deleted from device history",
        )

    db.tickets.update_one(
        {"id": ticket["id"]},
        {"$set": {
            "is_deleted": True,
            "deleted_at": now_iso(),
            "deleted_by": user["id"],
            "updated_at": now_iso(),
        }},
    )
    return {"ok": True, "ticket_id": ticket["id"], "is_deleted": True}

@api.post("/device-history/{ticket_id}/restore")
async def restore_device_history(ticket_id: str, user=Depends(require_sub_admin)):
    """Restore a soft-deleted device-history entry."""
    ticket = db.tickets.find_one({"id": ticket_id}) or db.tickets.find_one({"ticket_no": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Device history record not found")

    if user.get("role") == "sub_admin":
        company_ids = user.get("assigned_company_ids") or []
        if ticket.get("company_id") not in company_ids:
            raise HTTPException(status_code=403, detail="Not authorized to restore this record")

    db.tickets.update_one(
        {"id": ticket["id"]},
        {"$set": {"is_deleted": False, "updated_at": now_iso()},
         "$unset": {"deleted_at": "", "deleted_by": ""}},
    )
    return {"ok": True, "ticket_id": ticket["id"], "is_deleted": False}

@api.get("/devices/{device_id}")
async def get_device(device_id: str, user=Depends(get_current_user)):
    device = db.devices.find_one({"device_id": device_id}, {"_id": 0})
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    tickets = list(db.tickets.find(
        {"$or": [{"device_id": device_id}, {"device_ids": device_id}]},
        {"_id": 0}
    ).sort("created_at", -1).limit(100))
    for t in tickets:
        if t.get("assigned_engineer_id"):
            eng = db.users.find_one(
                {"id": t["assigned_engineer_id"]}, {"_id": 0, "name": 1}
            )
            t["engineer_name"] = eng["name"] if eng else None
    return {"device": device, "history": tickets}

# ---------- TICKETS ----------
def _get_or_create_device(company: dict, d: DeviceCreate) -> dict:
    serial = (d.serial_number or "").strip() or None
    existing = None
    if serial:
        # Serial number is globally unique across all companies
        existing = db.devices.find_one({"serial_number": serial})
        if existing and existing.get("company_id") != company["id"]:
            raise HTTPException(
                status_code=400,
                detail=f"Serial number '{serial}' is already registered under a different company."
            )
    if existing:
        updates = {}
        if d.warranty_status != existing.get("warranty_status"):
            updates["warranty_status"] = d.warranty_status
        if d.warranty_expiry and d.warranty_expiry != existing.get("warranty_expiry"):
            updates["warranty_expiry"] = d.warranty_expiry
        if updates:
            updates["updated_at"] = now_iso()
            db.devices.update_one({"device_id": existing["device_id"]},
                                  {"$set": updates})
            existing.update(updates)
        existing.pop("_id", None)
        return existing
    dev_id = next_device_id()
    doc = {
        "id": new_id(),
        "device_id": dev_id,
        "company_id": company["id"],
        "company_name": company["company_name"],
        "brand": d.brand,
        "model": d.model,
        "device_name": d.device_name or f"{d.brand} {d.model}",
        "device_type": d.device_type,
        "warranty_status": d.warranty_status,
        "warranty_expiry": d.warranty_expiry,
        "purchase_date": d.purchase_date,
        "notes": d.notes,
        "is_deleted": False,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    if serial:
        doc["serial_number"] = serial
    db.devices.insert_one(doc)
    doc.pop("_id", None)
    return doc

def _ticket_device_ids(ticket: dict) -> List[str]:
    ids: List[str] = []
    raw_ids = ticket.get("device_ids")
    if isinstance(raw_ids, list):
        ids.extend([dev_id for dev_id in raw_ids if dev_id])
    if ticket.get("device_id"):
        ids.append(ticket["device_id"])
    seen = set()
    ordered_ids = []
    for dev_id in ids:
        if dev_id not in seen:
            ordered_ids.append(dev_id)
            seen.add(dev_id)
    return ordered_ids

def _fetch_ticket_devices(ticket: dict) -> List[dict]:
    device_ids = _ticket_device_ids(ticket)
    if not device_ids:
        return []
    devices_map = {
        d["device_id"]: d
        for d in db.devices.find({"device_id": {"$in": device_ids}}, {"_id": 0})
    }
    return [devices_map[dev_id] for dev_id in device_ids if dev_id in devices_map]

def _attach_ticket_devices(ticket: dict, devices_map: Dict[str, dict]) -> None:
    devices = [
        devices_map[dev_id]
        for dev_id in _ticket_device_ids(ticket)
        if dev_id in devices_map
    ]
    if devices:
        ticket["devices"] = devices
        ticket["device"] = devices[0]
        ticket["device_count"] = len(devices)

def _log_status(ticket_id: str, actor: dict, old_status: Optional[str],
                new_status: str, remarks: Optional[str] = None):
    db.ticket_status_logs.insert_one({
        "id": new_id(),
        "ticket_id": ticket_id,
        "old_status": old_status,
        "new_status": new_status,
        "changed_by": actor["id"],
        "changed_by_name": actor.get("name"),
        "changed_by_role": actor.get("role"),
        "remarks": remarks,
        "timestamp": now_iso(),
    })

def _ticket_full(ticket: dict) -> dict:
    if not ticket:
        return ticket
    ticket = clean(ticket)

    device_ids = _ticket_device_ids(ticket)
    company_ids = [ticket.get("company_id")] if ticket.get("company_id") else []
    user_ids = []
    if ticket.get("assigned_engineer_id"):
        user_ids.append(ticket["assigned_engineer_id"])
    if ticket.get("created_by"):
        user_ids.append(ticket["created_by"])

    if device_ids:
        devices_map = {
            d["device_id"]: d
            for d in db.devices.find({"device_id": {"$in": device_ids}}, {"_id": 0})
        }
        _attach_ticket_devices(ticket, devices_map)

    if company_ids:
        companies = list(db.companies.find({"id": {"$in": company_ids}}, {"_id": 0}))
        if companies:
            ticket["company"] = companies[0]

    if user_ids:
        users = {u["id"]: u for u in db.users.find(
            {"id": {"$in": user_ids}},
            {"_id": 0, "password_hash": 0}
        )}
        if ticket.get("assigned_engineer_id") in users:
            ticket["engineer"] = users[ticket["assigned_engineer_id"]]
        if ticket.get("created_by") in users:
            ticket["created_by_user"] = users[ticket["created_by"]]

    return ticket

@api.post("/tickets")
async def create_ticket(payload: TicketCreate, admin=Depends(require_ticket_operator)):
    company = db.companies.find_one({"id": payload.company_id})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    if company.get("status") == "inactive":
        raise HTTPException(status_code=400, detail="Company is inactive")

    requested_devices = payload.devices or ([payload.device] if payload.device else [])
    if not requested_devices:
        raise HTTPException(status_code=400, detail="At least one device is required")

    devices = []
    seen_device_ids = set()
    for requested_device in requested_devices:
        device_doc = _get_or_create_device(company, requested_device)
        if device_doc["device_id"] in seen_device_ids:
            continue
        devices.append(device_doc)
        seen_device_ids.add(device_doc["device_id"])
    if not devices:
        raise HTTPException(status_code=400, detail="At least one device is required")

    primary_device = devices[0]
    device_ids = [device["device_id"] for device in devices]
    serial_numbers = [
        device.get("serial_number")
        for device in devices
        if device.get("serial_number")
    ]
    ticket_no = next_ticket_number()
    ticket = {
        "id": new_id(),
        "ticket_no": ticket_no,
        "ticket_number": ticket_no,
        "company_id": company["id"],
        "company_name": company["company_name"],
        "customer_name": payload.customer_name or company.get("contact_person"),
        "customer_phone": payload.customer_phone or company.get("phone"),
        "customer_email": payload.customer_email,
        "customer_company": company["company_name"],
        "customer_address": company.get("address"),
        "current_address": payload.current_address,
        "company_address": company.get("address"),
        "company_city": company.get("city"),
        "company_state": company.get("state"),
        "company_pincode": company.get("pincode"),
        "contact_person": company.get("contact_person"),
        "contact_source": payload.contact_source,
        "issue_description": payload.issue_description,
        "problem_description": payload.issue_description,
        "priority": payload.priority,
        "product_reference_number": (payload.product_reference_number or None),
        "oem_reference_number": (payload.oem_reference_number or None),
        "device_id": primary_device["device_id"],
        "device_ids": device_ids,
        "device_count": len(devices),
        "device_name": primary_device.get("device_name"),
        "device_names": [device.get("device_name") for device in devices],
        "device_brand": primary_device.get("brand"),
        "device_model": primary_device.get("model"),
        "device_serial": primary_device.get("serial_number"),
        "serial_number": primary_device.get("serial_number"),
        "serial_numbers": serial_numbers,
        "status": "open",
        "assigned_engineer_id": None,
        "assigned_engineer_name": None,
        "engineer_notes": None,
        "admin_notes": None,
        "created_by": admin["id"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "completed_at": None,
        "engineer_location": None,
        "approved": False,
        "approved_at": None,
        "is_deleted": False,
        "report_id": None,
        "pdf_path": None,
    }
    db.tickets.insert_one(ticket)
    _log_status(ticket["id"], admin, None, "open",
                f"Ticket {ticket_no} created for {company['company_name']}")
    return _ticket_full(ticket)

@api.get("/tickets")
async def list_tickets(
    status: Optional[str] = None,
    company_id: Optional[str] = None,
    mine: bool = False,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    user=Depends(get_current_user),
):
    q: Dict[str, Any] = {"is_deleted": {"$ne": True}}
    if status:
        q["status"] = status
    if company_id:
        q["company_id"] = company_id
    if user["role"] == "engineer" or mine:
        q["assigned_engineer_id"] = user["id"]
    if user["role"] == "sub_admin":
        q.update(_sub_admin_ticket_scope(user))
    if user["role"] == "ticket_admin":
        if status in ("closed", "rejected"):
            q["status"] = "__not_allowed__"
        elif not status:
            q["status"] = {"$nin": ["closed", "rejected"]}

    # company_admin: scope to own company only
    if user["role"] == "company_admin":
        q["company_id"] = user.get("company_id")

    skip = (page - 1) * per_page

    # Single aggregation instead of 2 separate queries
    pipeline = [
        {"$match": q},
        {
            "$facet": {
                "count": [{"$count": "total"}],
                "items": [
                    {"$sort": {"created_at": -1}},
                    {"$skip": skip},
                    {"$limit": per_page},
                    {"$project": {
                        "_id": 0,
                        "id": 1,
                        "device_id": 1,
                        "device_ids": 1,
                        "device_count": 1,
                        "ticket_no": 1,
                        "ticket_number": 1,
                        "customer_name": 1,
                        "customer_phone": 1,
                        "customer_email": 1,
                        "customer_company": 1,
                        "created_by": 1,
                        "company_id": 1,
                        "company_name": 1,
                        "company_address": 1,
                        "company_city": 1,
                        "company_state": 1,
                        "company_pincode": 1,
                        "contact_person": 1,
                        "current_address": 1,
                        "customer_address": 1,
                        "assigned_engineer_name": 1,
                        "assigned_engineer_id": 1,
                        "status": 1,
                        "priority": 1,
                        "issue_description": 1,
                        "created_at": 1,
                        "approved_at": 1,
                        "completed_at": 1,
                        "product_reference_number": 1,
                        "oem_reference_number": 1,
                        "is_deleted": 1,
                        "has_issue": 1,
                    }},
                ]
            }
        }
    ]
    result = list(db.tickets.aggregate(pipeline))
    if result:
        r = result[0]
        total = r["count"][0]["total"] if r["count"] else 0
        tickets = r["items"]
    else:
        total = 0
        tickets = []

    device_ids = list({
        dev_id
        for t in tickets
        for dev_id in _ticket_device_ids(t)
    })
    user_ids = list({
        user_id
        for t in tickets
        for user_id in (t.get("assigned_engineer_id"), t.get("created_by"))
        if user_id
    })

    devices_map = {d["device_id"]: d for d in db.devices.find(
        {"device_id": {"$in": device_ids}},
        {
            "_id": 0,
            "brand": 1,
            "model": 1,
            "device_id": 1,
            "warranty_status": 1,
            "warranty_expiry": 1,
            "device_name": 1,
            "serial_number": 1,
        }
    )} if device_ids else {}

    users_map = {e["id"]: e for e in db.users.find(
        {"id": {"$in": user_ids}},
        {"_id": 0, "name": 1, "id": 1, "role": 1, "is_remote": 1}
    )} if user_ids else {}

    for t in tickets:
        _attach_ticket_devices(t, devices_map)
        if t.get("assigned_engineer_id"):
            t["engineer"] = users_map.get(t["assigned_engineer_id"])
        if t.get("created_by"):
            t["created_by_user"] = users_map.get(t["created_by"])

    return {"items": tickets, "total": total, "page": page, "per_page": per_page}

@api.get("/tickets/{ticket_id}")
async def get_ticket(ticket_id: str, user=Depends(get_current_user)):
    ticket = db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if user["role"] == "engineer" and ticket.get("assigned_engineer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Not your ticket")
    full = _ticket_full(ticket)
    full["status_logs"] = list(db.ticket_status_logs.find(
        {"ticket_id": ticket_id}, {"_id": 0}
    ).sort("timestamp", -1).limit(200))
    full["activity"] = [
        {"id": s["id"], "ticket_id": s["ticket_id"],
         "action": f"status_{s['new_status']}",
         "actor_id": s.get("changed_by"),
         "actor_name": s.get("changed_by_name"),
         "actor_role": s.get("changed_by_role"),
         "details": s.get("remarks"),
         "timestamp": s["timestamp"]}
        for s in full["status_logs"]
    ]
    if full.get("report_id"):
        rep = db.service_reports.find_one({"id": full["report_id"]}, {"_id": 0})
        full["report"] = rep
    full_device_ids = _ticket_device_ids(full)
    if full_device_ids:
        history = list(db.tickets.find(
            {
                "id": {"$ne": ticket_id},
                "$or": [
                    {"device_id": {"$in": full_device_ids}},
                    {"device_ids": {"$in": full_device_ids}},
                ],
            },
            {"_id": 0, "ticket_no": 1, "ticket_number": 1, "status": 1,
             "created_at": 1, "issue_description": 1, "problem_description": 1,
             "assigned_engineer_id": 1, "device_id": 1, "device_ids": 1}
        ).sort("created_at", -1).limit(20))
        full["device_history"] = history
    return full

@api.post("/tickets/{ticket_id}/assign")
async def assign_ticket(ticket_id: str, payload: TicketAssign,
                        admin=Depends(require_ticket_operator)):
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    old_status = ticket["status"]

    if admin.get("role") == "ticket_admin" and ticket.get("status") in ("closed", "rejected"):
        raise HTTPException(status_code=403, detail="Ticket admins can only reassign active tickets")

    if payload.is_outsource:
        if admin.get("role") == "ticket_admin":
            raise HTTPException(status_code=403, detail="Ticket admins can only assign internal engineers")
        if not payload.outsource_name:
            raise HTTPException(status_code=400, detail="Outsource engineer name required")
        db.tickets.update_one({"id": ticket_id}, {"$set": {
            "assigned_engineer_id": None,
            "assigned_engineer_name": payload.outsource_name,
            "is_outsource": True,
            "outsource": {
                "name": payload.outsource_name,
                "location": payload.outsource_company,
                "phone": payload.outsource_phone,
                "price": payload.outsource_price,
                "notes": payload.outsource_notes,
            },
            "status": "assigned",
            "updated_at": now_iso(),
        }})
        _log_status(ticket_id, admin, old_status, "assigned",
                    f"Outsourced to {payload.outsource_name}")
    else:
        if not payload.engineer_id:
            raise HTTPException(status_code=400, detail="Engineer ID required")
        eng = db.users.find_one(
            {"id": payload.engineer_id, "role": "engineer", "is_active": True}
        )
        if not eng:
            raise HTTPException(status_code=404, detail="Engineer not found")
        db.tickets.update_one({"id": ticket_id}, {"$set": {
            "assigned_engineer_id": payload.engineer_id,
            "assigned_engineer_name": eng["name"],
            "is_outsource": False,
            "outsource": None,
            "status": "assigned",
            "updated_at": now_iso(),
        }})
        _log_status(ticket_id, admin, old_status, "assigned", f"Assigned to {eng['name']}")
        db.notifications.insert_one({
            "id": new_id(),
            "user_id": payload.engineer_id,
            "title": "New ticket assigned",
            "message": f"Ticket {ticket['ticket_no']} has been assigned to you",
            "type": "ticket_assigned",
            "ticket_id": ticket_id,
            "read_status": False,
            "read": False,
            "created_at": now_iso(),
        })

        # Push notification to engineer
        eng_user = db.users.find_one({"id": payload.engineer_id})
        if eng_user and eng_user.get("fcm_token"):
            send_push_notification(
                token=eng_user["fcm_token"],
                title="New Ticket Assigned",
                body=f"Ticket {ticket['ticket_no']} assigned to you",
                data={"ticket_id": ticket_id, "type": "ticket_assigned"}
            )
    return _ticket_full(db.tickets.find_one({"id": ticket_id}, {"_id": 0}))

@api.post("/tickets/{ticket_id}/outsource-complete")
async def outsource_complete(ticket_id: str, admin=Depends(require_ticket_operator)):
    """Mark an outsourced ticket as completed manually by admin/sub-admin."""
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if not ticket.get("is_outsource"):
        raise HTTPException(status_code=400, detail="Not an outsource ticket")
    db.tickets.update_one({"id": ticket_id}, {"$set": {
        "status": "closed",
        "completed_at": now_iso(),
        "updated_at": now_iso(),
    }})
    _log_status(ticket_id, admin, ticket["status"], "closed",
                "Outsource job marked complete manually")
    return _ticket_full(db.tickets.find_one({"id": ticket_id}, {"_id": 0}))

@api.get("/tickets/{ticket_id}/outsource-pdf")
async def outsource_internal_pdf(ticket_id: str, auth: Optional[str] = None,
                                  user=Depends(get_current_user)):
    """Generate and stream the internal outsource PDF for accounts team."""
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if not ticket.get("is_outsource"):
        raise HTTPException(status_code=400, detail="Not an outsource ticket")
    outsource = ticket.get("outsource") or {}
    creator = db.users.find_one({"id": ticket.get("created_by")}, {"_id": 0, "name": 1})
    created_by = creator.get("name", "") if creator else ""
    ticket_for_pdf = {**ticket, "devices": _fetch_ticket_devices(ticket)}
    pdf_bytes = build_outsource_internal_pdf(ticket_for_pdf, outsource, created_by)
    filename = f"Outsource-Internal-{ticket.get('ticket_no', ticket_id)}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api.post("/tickets/{ticket_id}/service-report")
async def create_service_report(ticket_id: str, payload: ServiceReportCreate,
                                 admin=Depends(require_ticket_operator)):
    """Generate a client-facing service report PDF for outsource or admin-closed tickets."""
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")

    devices = _fetch_ticket_devices(ticket)
    device = devices[0] if devices else {}
    company  = db.companies.find_one({"id": ticket.get("company_id")}, {"_id": 0}) if ticket.get("company_id") else None
    outsource = ticket.get("outsource") or {}

    if ticket.get("is_outsource"):
        engineer_info = {
            "name": outsource.get("name", ""),
            "email": outsource.get("company", ""),
            "phone": outsource.get("phone", ""),
            "skills": ["Outsource Partner"],
            "is_outsource": True,
            "outsource_company": outsource.get("location", outsource.get("company", "")),
            "outsource_price": outsource.get("price"),
        }
    else:
        engineer_info = db.users.find_one(
            {"id": ticket.get("assigned_engineer_id")},
            {"_id": 0, "password_hash": 0}
        ) or {}

    report_id = new_id()
    report = {
        "id": report_id,
        "ticket_id": ticket_id,
        "work_notes": payload.work_done,
        "resolution_summary": payload.resolution_summary or payload.work_done,
        "parts_used": payload.parts_used or [],
        "customer_signature": None,
        "customer_signed_name": payload.customer_signed_name or ticket.get("customer_name"),
        "signed_at": now_iso(),
        "outsource": outsource,
        "created_by_name": admin.get("name"),
        "photos_before": [],
        "photos_after": [],
    }

    ticket_for_pdf = {
        **ticket,
        "report_id": report_id,
        "company": company,
        "devices": devices,
    }
    pdf_bytes = build_service_report_pdf(
        ticket=ticket_for_pdf, device=device,
        engineer=engineer_info, report=report,
    )
    pdf_path = f"reports/{ticket.get('ticket_no', ticket_id)}-service-report.pdf"
    put_object(pdf_path, pdf_bytes, "application/pdf")

    db.service_reports.replace_one({"ticket_id": ticket_id}, report, upsert=True)
    db.tickets.update_one({"id": ticket_id}, {"$set": {
        "report_id": report_id,
        "pdf_path": pdf_path,
        "pdf_url": f"/api/files/{pdf_path}",
        "updated_at": now_iso(),
    }})
    return {"ok": True, "pdf_path": pdf_path}

class TicketAddressUpdate(BaseModel):
    current_address: str

@api.patch("/tickets/{ticket_id}/address")
async def update_ticket_address(
    ticket_id: str, payload: TicketAddressUpdate,
    user=Depends(require_ticket_operator),
):
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    db.tickets.update_one({"id": ticket_id}, {"$set": {
        "current_address": payload.current_address.strip(),
        "updated_at": now_iso(),
    }})
    _log_status(ticket_id, user, ticket["status"], ticket["status"],
                f"Address updated to: {payload.current_address.strip()}")
    return _ticket_full(db.tickets.find_one({"id": ticket_id}, {"_id": 0}))

@api.post("/tickets/{ticket_id}/status")
async def update_status(ticket_id: str, payload: StatusUpdate,
                        user=Depends(get_current_user)):
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if user["role"] == "engineer" and ticket.get("assigned_engineer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Not your ticket")

    if payload.status == "rejected":
        new_status = "open"
        db.tickets.update_one(
            {"id": ticket_id},
            {"$set": {
                "status": "open",
                "assigned_engineer_id": None,
                "assigned_engineer_name": None,
                "reject_reason": payload.reject_reason,
                "updated_at": now_iso(),
            }}
        )
        _log_status(ticket_id, user, ticket["status"], "rejected",
                    payload.reject_reason)
        db.notifications.insert_one({
            "id": new_id(), "user_id": "admin", "ticket_id": ticket_id,
            "title": f"Ticket {ticket['ticket_no']} rejected",
            "message": payload.reject_reason or "",
            "type": "status_rejected", "read_status": False, "read": False,
            "created_at": now_iso(),
        })
        return _ticket_full(db.tickets.find_one({"id": ticket_id}, {"_id": 0}))

    if payload.status == "report_generated":
        report = db.service_reports.find_one({"ticket_id": ticket_id})
        if not report:
            raise HTTPException(status_code=400,
                                detail="Cannot mark report_generated without a service report")

    updates = {"status": payload.status, "updated_at": now_iso()}
    if payload.status in ("closed", "report_generated"):
        updates["completed_at"] = now_iso()
    if payload.latitude is not None and payload.longitude is not None:
        updates["engineer_location"] = {
            "lat": payload.latitude, "lng": payload.longitude,
            "updated_at": now_iso(),
        }
    db.tickets.update_one({"id": ticket_id}, {"$set": updates})
    _log_status(ticket_id, user, ticket["status"], payload.status, payload.note)
    if payload.status in ("accepted", "resolved", "completed_with_signature",
                          "report_generated", "closed"):
        db.notifications.insert_one({
            "id": new_id(), "user_id": "admin", "ticket_id": ticket_id,
            "title": f"Ticket {ticket['ticket_no']} → {payload.status}",
            "message": payload.note or "",
            "type": f"status_{payload.status}",
            "read_status": False, "read": False, "created_at": now_iso(),
        })
    return _ticket_full(db.tickets.find_one({"id": ticket_id}, {"_id": 0}))

@api.post("/tickets/{ticket_id}/location")
async def update_location(ticket_id: str, payload: LocationUpdate,
                          user=Depends(require_engineer)):
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket or ticket.get("assigned_engineer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Not your ticket")
    db.tickets.update_one(
        {"id": ticket_id},
        {"$set": {"engineer_location": {
            "lat": payload.latitude, "lng": payload.longitude,
            "updated_at": now_iso()
        }}}
    )
    return {"ok": True}

@api.post("/tickets/{ticket_id}/report")
async def submit_report(ticket_id: str, payload: ReportSubmit,
                        user=Depends(require_engineer)):
    """
    Submit signed service report.
    Security: only assigned engineer; signature required.
    Flow: ticket → completed_with_signature → PDF generated → report_generated.
    """
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if ticket.get("assigned_engineer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Not your ticket")
    if not payload.customer_signature:
        raise HTTPException(status_code=400, detail="Customer signature is required")

    db.tickets.update_one({"id": ticket_id}, {
        "$set": {"status": "completed_with_signature", "updated_at": now_iso()}
    })
    _log_status(ticket_id, user, ticket["status"],
                "completed_with_signature",
                "Customer signature captured")

    devices = _fetch_ticket_devices(ticket)
    device = devices[0] if devices else {}
    company = db.companies.find_one(
        {"id": ticket.get("company_id")}, {"_id": 0}
    ) if ticket.get("company_id") else None
    engineer = db.users.find_one(
        {"id": user["id"]}, {"_id": 0, "password_hash": 0}
    )

    report_id = new_id()
    report = {
        "id": report_id,
        "ticket_id": ticket_id,
        "ticket_no": ticket["ticket_no"],
        "engineer_id": user["id"],
        "engineer_name": engineer["name"],
        "customer_name": payload.customer_signed_name or ticket.get("customer_name"),
        "engineer_notes": payload.engineer_notes,
        "resolution_summary": payload.resolution_summary or payload.engineer_notes,
        "parts_used": [p.model_dump() for p in payload.parts_used],
        "before_images": payload.before_images,
        "after_images": payload.after_images,
        "customer_signature": payload.customer_signature,
        "customer_signed_name": payload.customer_signed_name,
        "signed_at": now_iso(),
        "generated_at": now_iso(),
        "pdf_url": None,
        "pdf_path": None,
        "work_notes": payload.engineer_notes,
        "photos_before": payload.before_images,
        "photos_after": payload.after_images,
    }

    pdf_path = None
    try:
        ticket_for_pdf = {
            **ticket,
            "report_id": report_id,
            "company": company,
            "devices": devices,
        }
        pdf_bytes = build_service_report_pdf(
            ticket=ticket_for_pdf, device=device,
            engineer=engineer, report=report,
        )
        pdf_path = f"reports/{ticket['ticket_no']}-{report_id}.pdf"
        put_object(pdf_path, pdf_bytes, "application/pdf")
        db.attachments.insert_one({
            "id": new_id(),
            "ticket_id": ticket_id,
            "uploaded_by": user["id"],
            "file_name": f"{ticket['ticket_no']}.pdf",
            "file_type": "application/pdf",
            "file_url": f"/api/files/{pdf_path}",
            "storage_path": pdf_path,
            "size": len(pdf_bytes),
            "is_deleted": False,
            "uploaded_at": now_iso(),
        })
        report["pdf_url"] = f"/api/files/{pdf_path}"
        report["pdf_path"] = pdf_path
    except Exception as e:
        logger.error(f"PDF generation/upload failed: {e}")

    db.service_reports.replace_one(
        {"ticket_id": ticket_id}, report, upsert=True
    )

    next_status = "report_generated" if pdf_path else "completed_with_signature"
    db.tickets.update_one(
        {"id": ticket_id},
        {"$set": {
            "report_id": report_id,
            "status": next_status,
            "pdf_path": pdf_path,
            "pdf_url": report.get("pdf_url"),
            "updated_at": now_iso(),
            "completed_at": now_iso() if next_status == "report_generated" else None,
        }}
    )
    _log_status(ticket_id, user, "completed_with_signature",
                next_status, "Service report generated")
    db.notifications.insert_one({
        "id": new_id(), "user_id": "admin", "ticket_id": ticket_id,
        "title": f"Service report ready: {ticket['ticket_no']}",
        "message": "Customer-signed PDF is ready for review.",
        "type": "report_ready", "read_status": False, "read": False,
        "created_at": now_iso(),
    })

    # Push to admins and sub_admins
    admins = db.users.find({"role": {"$in": ["admin", "sub_admin"]}, "fcm_token": {"$exists": True}})
    for a in admins:
        send_push_notification(
            token=a["fcm_token"],
            title="Service Report Ready",
            body=f"Ticket {ticket['ticket_no']} report is ready for review",
            data={"ticket_id": ticket_id, "type": "report_ready"}
        )

        
    # Send email to client with PDF
    try:
        company = db.companies.find_one({"id": ticket.get("company_id")}, {"_id": 0}) if ticket.get("company_id") else None
        client_email = (company or {}).get("client_email") or (company or {}).get("email")
        if client_email and pdf_path:
            pdf_bytes_for_email = get_object(pdf_path)
            send_ticket_email(
                client_email, ticket, pdf_bytes_for_email,
                subject_prefix="Report Ready: ",
                sender_name=engineer.get("name") if engineer else None,
                sender_email=engineer.get("email") if engineer else None,
            )
    except Exception as e:
        logger.error(f"Email on report generation failed: {e}")

    report.pop("_id", None)
    return report

@api.post("/tickets/{ticket_id}/approve")
async def approve_ticket(ticket_id: str, admin=Depends(require_sub_admin)):
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    approved_at = now_iso()
    db.tickets.update_one(
        {"id": ticket_id},
        {"$set": {
            "approved": True,
            "approved_at": approved_at,
            "approved_by": admin["id"],
            "approved_by_name": admin.get("name"),
            "status": "closed",
            "updated_at": approved_at,
            "completed_at": approved_at,
        }}
    )
    _log_status(ticket_id, admin, ticket["status"], "closed",
                "Service report approved and ticket closed")
    
    # Push to all admins and sub_admins
    admins = db.users.find({"role": {"$in": ["admin", "sub_admin"]}, "fcm_token": {"$exists": True}})
    for a in admins:
        send_push_notification(
            token=a["fcm_token"],
            title="Ticket Closed",
            body=f"Ticket {ticket.get('ticket_no')} has been closed successfully",
            data={"ticket_id": ticket_id, "type": "ticket_closed"}
        )
        

    # Send 'Ticket Closed' email to customer with engineer PDF report attached.
    email_result = None
    try:
        updated_ticket = db.tickets.find_one({"id": ticket_id}, {"_id": 0}) or ticket
        # Prefer ticket.customer_email (Feature 1), fall back to company email.
        customer_email = updated_ticket.get("customer_email")
        if not customer_email:
            company = db.companies.find_one(
                {"id": updated_ticket.get("company_id")}, {"_id": 0}
            ) if updated_ticket.get("company_id") else None
            customer_email = (company or {}).get("client_email") or (company or {}).get("email")

        pdf_path = updated_ticket.get("pdf_path")
        pdf_bytes = None
        if pdf_path:
            try:
                data, _ = get_object(pdf_path)
                pdf_bytes = data
            except Exception as e:
                logger.error(f"Could not load PDF {pdf_path} for closure email: {e}")

        if customer_email:
            email_result = send_ticket_closed_email(
                customer_email, updated_ticket, pdf_bytes,
                sender_name=admin.get("name"),
                sender_email=admin.get("email"),
            )
            db.tickets.update_one(
                {"id": ticket_id},
                {"$set": {
                    "closure_email_sent_to": customer_email,
                    "closure_email_sent_at": now_iso(),
                    "closure_email_status": email_result,
                }},
            )
        else:
            logger.warning(f"No customer_email on ticket {ticket.get('ticket_no')} - closure email skipped")
    except Exception as e:
        logger.error(f"Email on ticket close failed: {e}")

    return _ticket_full(db.tickets.find_one({"id": ticket_id}, {"_id": 0}))

@api.post("/tickets/{ticket_id}/notify-customer")
async def notify_customer(ticket_id: str, user=Depends(require_sub_admin)):
    """Manually send a status notification email to the customer."""
    ticket = db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    company = db.companies.find_one({"id": ticket.get("company_id")}, {"_id": 0}) if ticket.get("company_id") else None
    client_email = (company or {}).get("client_email") or (company or {}).get("email")
    if not client_email:
        raise HTTPException(status_code=400, detail="No client email set for this company")
    pdf_path = ticket.get("pdf_path")
    pdf_bytes = get_object(pdf_path) if pdf_path else None
    send_ticket_email(
        client_email, ticket, pdf_bytes,
        subject_prefix="Update: ",
        sender_name=user.get("name"),
        sender_email=user.get("email"),
    )
    return {"ok": True, "sent_to": client_email}

# ---------- REPORTS ----------
@api.get("/reports/{ticket_id}")
async def get_report(ticket_id: str, user=Depends(get_current_user)):
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if user["role"] == "engineer" and ticket.get("assigned_engineer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Not your ticket")
    report = db.service_reports.find_one({"ticket_id": ticket_id}, {"_id": 0})
    if not report:
        raise HTTPException(status_code=404, detail="No report for this ticket")
    return report

# ---------- FILES (local storage server) ----------
@api.get("/files/{path:path}")
async def serve_file(path: str, request: Request):
    """Serve files from local storage. Auth via header or ?auth=token."""
    from auth import extract_token
    token = extract_token(request)
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    payload = decode_token(token)
    user = db.users.find_one({"id": payload["sub"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")

    attachment = db.attachments.find_one({
        "storage_path": path, "is_deleted": False
    })
    if attachment:
        if user["role"] == "engineer":
            ticket = db.tickets.find_one({"id": attachment["ticket_id"]})
            if not ticket or ticket.get("assigned_engineer_id") != user["id"]:
                raise HTTPException(status_code=403, detail="Access denied")
    try:
        data, _ = get_object(path)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="File not found")
    content_type = "application/pdf" if path.endswith(".pdf") else \
                   "image/jpeg" if path.endswith((".jpg", ".jpeg")) else \
                   "image/png" if path.endswith(".png") else \
                   "application/octet-stream"
    filename = os.path.basename(path)
    return Response(content=data, media_type=content_type, headers={
        "Content-Disposition": f'inline; filename="{filename}"'
    })

# Back-compat endpoint
@api.get("/tickets/{ticket_id}/pdf")
async def get_ticket_pdf(ticket_id: str, request: Request):
    from auth import extract_token
    token = extract_token(request)
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    payload = decode_token(token)
    user = db.users.find_one({"id": payload["sub"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if user["role"] == "engineer" and ticket.get("assigned_engineer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Not your ticket")
    if not ticket.get("pdf_path"):
        raise HTTPException(status_code=404, detail="PDF not generated yet")
    try:
        data, _ = get_object(ticket["pdf_path"])
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="PDF file missing")
    return Response(content=data, media_type="application/pdf", headers={
        "Content-Disposition": f"inline; filename=\"{ticket['ticket_no']}.pdf\""
    })

# ---------- NOTIFICATIONS ----------
@api.get("/notifications")
async def list_notifications(user=Depends(get_current_user)):
    target = "admin" if user["role"] == "admin" else user["id"]
    notes = list(db.notifications.find(
        {"user_id": target}, {"_id": 0}
    ).sort("created_at", -1).limit(50))
    for n in notes:
        if "read" not in n:
            n["read"] = n.get("read_status", False)
    return notes

@api.post("/notifications/{nid}/read")
async def mark_read(nid: str, user=Depends(get_current_user)):
    db.notifications.update_one(
        {"id": nid}, {"$set": {"read": True, "read_status": True}}
    )
    return {"ok": True}

# ---------- ATTENDANCE ----------
@api.post("/attendance/check-in")
async def check_in(payload: AttendanceAction, user=Depends(require_engineer)):
    today = date.today().isoformat()
    existing = db.attendance.find_one({"engineer_id": user["id"], "date": today})
    if existing and existing.get("check_in_time"):
        raise HTTPException(status_code=400, detail="Already checked in today")
    doc = {
        "id": new_id(),
        "engineer_id": user["id"],
        "date": today,
        "check_in_time": now_iso(),
        "check_in_location": {"lat": payload.latitude, "lng": payload.longitude}
            if payload.latitude is not None else None,
        "check_out_time": None,
        "attendance_status": "present",
        "created_at": now_iso(),
    }
    if existing:
        db.attendance.update_one({"id": existing["id"]}, {"$set": doc})
        doc["id"] = existing["id"]
    else:
        db.attendance.insert_one(doc)
    doc["check_in"] = doc["check_in_time"]
    doc["check_out"] = doc["check_out_time"]
    doc.pop("_id", None)
    return doc

@api.post("/attendance/check-out")
async def check_out(payload: AttendanceAction, user=Depends(require_engineer)):
    today = date.today().isoformat()
    existing = db.attendance.find_one({"engineer_id": user["id"], "date": today})
    if not existing or not existing.get("check_in_time"):
        raise HTTPException(status_code=400, detail="Not checked in")
    if existing.get("check_out_time"):
        raise HTTPException(status_code=400, detail="Already checked out")
    db.attendance.update_one(
        {"id": existing["id"]},
        {"$set": {
            "check_out_time": now_iso(),
            "check_out_location": {"lat": payload.latitude, "lng": payload.longitude}
                if payload.latitude is not None else None,
        }}
    )
    doc = db.attendance.find_one({"id": existing["id"]}, {"_id": 0})
    doc["check_in"] = doc.get("check_in_time")
    doc["check_out"] = doc.get("check_out_time")
    return doc

@api.get("/attendance/today")
async def attendance_today(user=Depends(require_engineer)):
    today = date.today().isoformat()
    doc = db.attendance.find_one({"engineer_id": user["id"], "date": today},
                                 {"_id": 0})
    if doc:
        doc["check_in"] = doc.get("check_in_time")
        doc["check_out"] = doc.get("check_out_time")
    return doc or {}

@api.get("/attendance/history")
async def attendance_history(user=Depends(require_engineer)):
    docs = list(db.attendance.find(
        {"engineer_id": user["id"]}, {"_id": 0}
    ).sort("date", -1).limit(60))
    for d in docs:
        d["check_in"] = d.get("check_in_time")
        d["check_out"] = d.get("check_out_time")
    return docs

# ---------- DASHBOARD ----------
@api.get("/dashboard/admin", dependencies=[Depends(require_admin_console)])
async def admin_dashboard(user=Depends(get_current_user)):
    cache_key = _cache_key("dashboard-admin", user)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached
    try:
        # Scope by assigned companies for sub_admin
        assigned_company_ids = []
        company_scope = {}
        if user.get("role") == "sub_admin":
            assigned_company_ids = user.get("assigned_company_ids") or []
            company_scope = _sub_admin_ticket_scope(user)
        ticket_scope = {"is_deleted": {"$ne": True}, **company_scope}
        if user.get("role") == "ticket_admin":
            ticket_scope["status"] = {"$nin": ["closed", "rejected"]}

        # ---- Ticket counts (ONE QUERY instead of MANY) ----
        ticket_counts_raw = list(db.tickets.aggregate([
            {"$match": ticket_scope},
            {
                "$group": {
                    "_id": "$status",
                    "count": {"$sum": 1}
                }
            }
        ]))

        counts = {s: 0 for s in TICKET_STATUSES}

        for item in ticket_counts_raw:
            counts[item["_id"]] = item["count"]

        counts["total"] = sum(counts.values())
        counts["completed"] = sum(
            counts.get(s, 0) for s in [
                "completed_with_signature", "report_generated", "closed"
            ]
        )
        counts["active"] = sum(
            counts.get(s, 0) for s in [
                "open", "assigned", "accepted", "travelling",
                "reached_site", "in_progress", "resolved",
                "completed_with_signature", "report_generated"
            ]
        )

        # ---- Engineers ----
        total_eng = db.users.count_documents({
            "role": "engineer",
            "is_active": True
        })

        available = db.users.count_documents({
            "role": "engineer",
            "is_active": True,
            "is_available": True
        })

        remote_eng = db.users.count_documents({
            "role": "engineer",
            "is_active": True,
            "is_remote": True
        })

        engineer_work_modes = list(db.users.find(
            {"role": "engineer", "is_active": True},
            {
                "_id": 0,
                "id": 1,
                "name": 1,
                "email": 1,
                "phone": 1,
                "is_remote": 1,
                "is_available": 1,
            },
        ).sort("name", 1).limit(100))

        # ---- Companies ----
        company_query = {"status": "active"}
        if user.get("role") == "sub_admin":
            company_query["id"] = {"$in": assigned_company_ids}
        total_co = db.companies.count_documents(company_query)

        # ---- Recent activity (LIMITED) ----
        if user.get("role") in ("sub_admin", "ticket_admin"):
            # Limit to recent 500 ticket IDs to avoid huge $in queries
            ticket_ids = [t["id"] for t in db.tickets.find(
                ticket_scope, {"_id": 0, "id": 1}
            ).sort("created_at", -1).limit(500)]
            log_query = {"ticket_id": {"$in": ticket_ids}}
            if user.get("role") == "sub_admin":
                log_query = {"$or": [log_query, {"target_user_id": user["id"]}]}
            logs = list(
                db.ticket_status_logs
                .find(log_query, {"_id": 0})
                .sort("timestamp", -1)
                .limit(5)
            )
        else:
            logs = list(
                db.ticket_status_logs
                .find({}, {"_id": 0})
                .sort("timestamp", -1)
                .limit(5)
            )

        recent = []
        for l in logs:
            if l.get("actor_type") == "sub_admin_assignment":
                recent.append({
                    "id": l["id"],
                    "actor_id": l.get("changed_by"),
                    "actor_name": l.get("changed_by_name"),
                    "actor_role": l.get("changed_by_role"),
                    "action": "sub_admin_companies_updated",
                    "details": f"Updated company assignments from {len(l.get('old_companies', []))} to {len(l.get('new_companies', []))}",
                    "timestamp": l["timestamp"],
                })
            else:
                recent.append({
                    "id": l["id"],
                    "actor_id": l.get("changed_by"),
                    "actor_name": l.get("changed_by_name"),
                    "actor_role": l.get("changed_by_role"),
                    "action": f"status_{l['new_status']}",
                    "details": l.get("remarks"),
                    "timestamp": l["timestamp"],
                })

        # ---- Warranty alerts (DB filter, not Python) ----
        today = date.today()
        in_30 = (today + timedelta(days=30)).isoformat()

        warranty_query = {
            "warranty_status": "active",
            "warranty_expiry": {
                "$gte": today.isoformat(),
                "$lte": in_30
            }
        }
        if user.get("role") == "sub_admin":
            warranty_query["company_id"] = {"$in": assigned_company_ids}

        warranty_alerts = list(
            db.devices.find(warranty_query, {"_id": 0}).limit(20)
        )

        response = {
            "ticket_counts": counts,
            "engineers": {
                "total": total_eng,
                "available": available,
                "remote": remote_eng,
                "work_modes": engineer_work_modes,
            },
            "companies": {
                "active": total_co
            },
            "recent_activity": recent,
            "warranty_alerts": warranty_alerts,
        }
        return _cache_set(cache_key, response, 30)

    except Exception as e:
        print("DASHBOARD ERROR:", str(e))
        raise HTTPException(status_code=500, detail="Dashboard error")

@api.get("/dashboard/engineer")
async def engineer_dashboard(user=Depends(require_engineer)):
    cache_key = _cache_key("dashboard-engineer", user)
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached
    eid = user["id"]
    # Single aggregation instead of 4 count_documents calls
    base_match = {"assigned_engineer_id": eid, "is_deleted": {"$ne": True}}
    pipeline = [
        {"$match": base_match},
        {
            "$facet": {
                "assigned": [{"$match": {"status": {"$in": ["assigned", "accepted"]}}}, {"$count": "count"}],
                "in_progress": [{"$match": {"status": {"$in": ["travelling", "reached_site", "in_progress"]}}}, {"$count": "count"}],
                "completed": [{"$match": {"status": {"$in": ["closed", "report_generated", "completed_with_signature"]}}}, {"$count": "count"}],
                "resolved": [{"$match": {"status": "resolved"}}, {"$count": "count"}]
            }
        }
    ]
    result = list(db.tickets.aggregate(pipeline))
    active_statuses = ["assigned", "accepted", "travelling", "reached_site", "in_progress"]
    active_tickets = list(db.tickets.find(
        {**base_match, "status": {"$in": active_statuses}},
        {
            "_id": 0,
            "id": 1,
            "ticket_no": 1,
            "ticket_number": 1,
            "customer_name": 1,
            "status": 1,
            "created_at": 1,
            "device_id": 1,
            "device_ids": 1,
            "device_count": 1,
        },
    ).sort("created_at", -1).limit(20))

    device_ids = list({
        dev_id
        for t in active_tickets
        for dev_id in _ticket_device_ids(t)
    })
    devices_map = {d["device_id"]: d for d in db.devices.find(
        {"device_id": {"$in": device_ids}},
        {"_id": 0, "brand": 1, "model": 1, "device_id": 1}
    )} if device_ids else {}
    for ticket in active_tickets:
        _attach_ticket_devices(ticket, devices_map)

    if result:
        r = result[0]
        response = {
            "assigned": r["assigned"][0]["count"] if r["assigned"] else 0,
            "in_progress": r["in_progress"][0]["count"] if r["in_progress"] else 0,
            "completed": r["completed"][0]["count"] if r["completed"] else 0,
            "resolved": r["resolved"][0]["count"] if r["resolved"] else 0,
            "is_remote": bool(user.get("is_remote", False)),
            "active_tickets": active_tickets,
        }
        return _cache_set(cache_key, response, 10)
    response = {
        "assigned": 0,
        "in_progress": 0,
        "completed": 0,
        "resolved": 0,
        "is_remote": bool(user.get("is_remote", False)),
        "active_tickets": active_tickets,
    }
    return _cache_set(cache_key, response, 10)

# ---------- ANALYTICS ----------
@api.get("/analytics", dependencies=[Depends(require_sub_admin)])
async def analytics(user=Depends(get_current_user)):
    today = date.today()
    cache_key = _cache_key("analytics", user, today.isoformat())
    cached = _cache_get(cache_key)
    if cached is not None:
        return cached

    # Scope by assigned companies for sub_admin
    scope = {}
    if user.get("role") == "sub_admin":
        scope = _sub_admin_ticket_scope(user)

    # 1. Per-day: single aggregation instead of 14 count_documents
    start_14 = (today - timedelta(days=13)).isoformat()
    per_day_raw = list(db.tickets.aggregate([
        {"$match": {"created_at": {"$gte": start_14}, **scope}},
        {"$group": {"_id": {"$substr": ["$created_at", 0, 10]}, "count": {"$sum": 1}}},
    ]))
    per_day_map = {r["_id"]: r["count"] for r in per_day_raw}
    per_day = [
        {"date": (today - timedelta(days=i)).isoformat(),
         "count": per_day_map.get((today - timedelta(days=i)).isoformat(), 0)}
        for i in range(13, -1, -1)
    ]

    # 2. Engineer performance: single aggregation instead of 2 queries per engineer
    eng_perf_raw = list(db.tickets.aggregate([
        {"$match": {"assigned_engineer_id": {"$ne": None}, **scope}},
        {"$group": {
            "_id": "$assigned_engineer_id",
            "completed": {"$sum": {"$cond": [{"$in": ["$status", ["closed", "report_generated"]]}, 1, 0]}},
            "active":    {"$sum": {"$cond": [{"$in": ["$status", ["closed", "rejected", "report_generated"]]}, 0, 1]}},
        }},
    ]))
    eng_name_map = {e["id"]: e["name"] for e in db.users.find(
        {"role": "engineer"}, {"_id": 0, "id": 1, "name": 1}
    )}
    perf = sorted([
        {"name": eng_name_map.get(r["_id"], r["_id"]),
         "completed": r["completed"], "active": r["active"]}
        for r in eng_perf_raw if r["_id"] in eng_name_map
    ], key=lambda x: -x["completed"])

    # 3. Repeat complaints: $lookup instead of per-row find_one
    repeat_pipeline = []
    if scope: repeat_pipeline.append({"$match": scope})
    repeat_pipeline += [
        {"$group": {"_id": "$device_id", "count": {"$sum": 1}}},
        {"$match": {"count": {"$gt": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10},
        {"$lookup": {"from": "devices", "localField": "_id", "foreignField": "device_id", "as": "device"}},
        {"$unwind": {"path": "$device", "preserveNullAndEmptyArrays": True}},
        {"$project": {"_id": 0, "device_id": "$_id", "count": 1,
                      "brand": "$device.brand", "model": "$device.model"}},
    ]
    repeats = list(db.tickets.aggregate(repeat_pipeline))

    # 4. Brand trend: $lookup instead of per-row find_one
    brand_pipeline = []
    if scope: brand_pipeline.append({"$match": scope})
    brand_pipeline += [
        {"$group": {"_id": "$device_id", "count": {"$sum": 1}}},
        {"$lookup": {"from": "devices", "localField": "_id", "foreignField": "device_id", "as": "device"}},
        {"$unwind": {"path": "$device", "preserveNullAndEmptyArrays": True}},
        {"$group": {"_id": {"$ifNull": ["$device.brand", "Unknown"]}, "tickets": {"$sum": "$count"}}},
        {"$sort": {"tickets": -1}},
        {"$limit": 8},
        {"$project": {"_id": 0, "brand": "$_id", "tickets": 1}},
    ]
    brand_trend = list(db.tickets.aggregate(brand_pipeline))

    # 5. Warranty alerts
    in_30 = (today + timedelta(days=30)).isoformat()
    warranty_alerts = list(db.devices.find(
        {"warranty_status": "active",
         "warranty_expiry": {"$gte": today.isoformat(), "$lte": in_30}},
        {"_id": 0}
    ).limit(20))

    # 6. Customer analytics - tickets per customer per month (last 6 months)
    start_6m = (today.replace(day=1) - timedelta(days=150)).isoformat()
    customer_raw = list(db.tickets.aggregate([
        {"$match": {"created_at": {"$gte": start_6m}, **scope}},
        {"$group": {
            "_id": {
                "customer": "$customer_name",
                "company": "$customer_company",
                "month": {"$substr": ["$created_at", 0, 7]}
            },
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id.month": -1, "count": -1}},
    ]))

    # Group by month
    customer_by_month = {}
    for r in customer_raw:
        month = r["_id"]["month"]
        if month not in customer_by_month:
            customer_by_month[month] = []
        customer_by_month[month].append({
            "customer": r["_id"]["customer"] or "Unknown",
            "company": r["_id"]["company"] or "",
            "count": r["count"]
        })

    # Top customers overall this month
    this_month = today.strftime("%Y-%m")
    top_customers = sorted(
        customer_by_month.get(this_month, []),
        key=lambda x: -x["count"]
    )[:10]

    # 7. Resolution time analytics — avg hours to close a ticket
    closed_tickets = list(db.tickets.find(
        {"status": {"$in": ["closed", "report_generated"]},
         "completed_at": {"$ne": None},
         "created_at": {"$ne": None},
         **scope},
        {"_id": 0, "created_at": 1, "completed_at": 1, "assigned_engineer_id": 1}
    ).limit(500))

    def hours_diff(start, end):
        try:
            from datetime import datetime as dt
            fmt = "%Y-%m-%dT%H:%M:%S"
            s = dt.fromisoformat(start[:19])
            e = dt.fromisoformat(end[:19])
            diff = (e - s).total_seconds() / 3600
            return round(diff, 1) if diff >= 0 else None
        except:
            return None

    resolution_hours = [
        h for t in closed_tickets
        if (h := hours_diff(t.get("created_at",""), t.get("completed_at",""))) is not None
    ]

    avg_hours = round(sum(resolution_hours) / len(resolution_hours), 1) if resolution_hours else 0
    min_hours = min(resolution_hours, default=0)
    max_hours = max(resolution_hours, default=0)

    # Bucket into ranges for chart
    buckets = [
        {"range": "< 4 hrs",   "min": 0,   "max": 4},
        {"range": "4–8 hrs",   "min": 4,   "max": 8},
        {"range": "8–24 hrs",  "min": 8,   "max": 24},
        {"range": "1–3 days",  "min": 24,  "max": 72},
        {"range": "3–7 days",  "min": 72,  "max": 168},
        {"range": "> 7 days",  "min": 168, "max": 999999},
    ]
    for b in buckets:
        b["count"] = sum(1 for h in resolution_hours if b["min"] <= h < b["max"])

    # Per engineer avg resolution
    eng_res_map = {}
    for t in closed_tickets:
        eid = t.get("assigned_engineer_id")
        if not eid: continue
        h = hours_diff(t.get("created_at",""), t.get("completed_at",""))
        if h is None: continue
        if eid not in eng_res_map: eng_res_map[eid] = []
        eng_res_map[eid].append(h)

    eng_name_map2 = {e["id"]: e["name"] for e in db.users.find(
        {"id": {"$in": list(eng_res_map.keys())}, "role": "engineer"},
        {"_id": 0, "id": 1, "name": 1},
    )}
    eng_resolution = sorted([
        {"name": eng_name_map2[eid], "avg_hours": round(sum(hrs) / len(hrs), 1), "tickets": len(hrs)}
        for eid, hrs in eng_res_map.items() if eid in eng_name_map2
    ], key=lambda x: x["avg_hours"])

    response = {
        "per_day": per_day,
        "engineer_performance": perf,
        "repeat_complaints": repeats,
        "brand_trend": brand_trend,
        "warranty_alerts": warranty_alerts,
        "customer_by_month": customer_by_month,
        "top_customers_this_month": top_customers,
        "resolution_time": {
            "avg_hours": avg_hours,
            "min_hours": min_hours,
            "max_hours": max_hours,
            "total_closed": len(resolution_hours),
            "buckets": buckets,
            "by_engineer": eng_resolution,
        },
    }
    return _cache_set(cache_key, response, 120)

@api.get("/live-locations")
async def live_locations(user=Depends(require_sub_admin)):
    q = {
        "engineer_location": {"$ne": None},
        "status": {"$in": ["travelling", "reached_site", "in_progress"]},
        **_sub_admin_ticket_scope(user),
    }
    tickets = list(db.tickets.find(
        q,
        {"_id": 0}
    ).limit(200))
    eng_ids = list({t.get("assigned_engineer_id") for t in tickets if t.get("assigned_engineer_id")})
    engineers = {e["id"]: e for e in db.users.find(
        {"id": {"$in": eng_ids}},
        {"_id": 0, "id": 1, "name": 1},
    )} if eng_ids else {}
    out = []
    for t in tickets:
        eng = engineers.get(t.get("assigned_engineer_id"))
        out.append({
            "ticket_id": t["id"],
            "ticket_number": t.get("ticket_no") or t.get("ticket_number"),
            "engineer_name": eng["name"] if eng else "Unknown",
            "status": t["status"],
            "location": t["engineer_location"],
            "customer_name": t.get("customer_name"),
        })
    return out

# ---------- TICKET ISSUES (Feature 4: Engineer Failure/Delay Reporting) ----------
class TicketIssueCreate(BaseModel):
    issue_type: Literal[
        "Customer Not Available", "Spare Part Pending", "OEM Support Pending",
        "Site Closed", "Material Not Received", "Network Issue",
        "Access Denied", "Other"
    ]
    description: str
    expected_resolution_date: Optional[str] = None
    priority: Literal["low", "medium", "high", "critical"] = "medium"

class TicketIssueUpdate(BaseModel):
    status: Optional[Literal["open", "resolved", "cancelled"]] = None
    description: Optional[str] = None
    expected_resolution_date: Optional[str] = None
    priority: Optional[Literal["low", "medium", "high", "critical"]] = None

@api.post("/tickets/{ticket_id}/issues")
async def report_ticket_issue(
    ticket_id: str, payload: TicketIssueCreate,
    user=Depends(require_engineer),
):
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if ticket.get("assigned_engineer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Not your ticket")

    issue_id = new_id()
    doc = {
        "id": issue_id,
        "ticket_id": ticket_id,
        "ticket_no": ticket.get("ticket_no") or ticket.get("ticket_number"),
        "engineer_id": user["id"],
        "engineer_name": user.get("name"),
        "company_id": ticket.get("company_id"),
        "issue_type": payload.issue_type,
        "description": payload.description,
        "expected_resolution_date": payload.expected_resolution_date,
        "priority": payload.priority,
        "status": "open",
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    db.ticket_issues.insert_one(doc)

    # Mark ticket has_issue flag and add status log entry
    db.tickets.update_one(
        {"id": ticket_id},
        {"$set": {"has_issue": True, "updated_at": now_iso()}}
    )
    _log_status(ticket_id, user, ticket["status"], ticket["status"],
                f"Issue reported: {payload.issue_type}")

    # Notify admins
    db.notifications.insert_one({
        "id": new_id(), "user_id": "admin", "ticket_id": ticket_id,
        "title": f"Issue reported on {ticket.get('ticket_no')}",
        "message": f"{user.get('name')}: {payload.issue_type}",
        "type": "issue_reported", "read_status": False, "read": False,
        "created_at": now_iso(),
    })
    doc.pop("_id", None)
    return doc

@api.get("/tickets/{ticket_id}/issues")
async def get_ticket_issues(ticket_id: str, user=Depends(get_current_user)):
    ticket = db.tickets.find_one({"id": ticket_id})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if user["role"] == "engineer" and ticket.get("assigned_engineer_id") != user["id"]:
        raise HTTPException(status_code=403, detail="Not your ticket")
    issues = list(db.ticket_issues.find({"ticket_id": ticket_id}, {"_id": 0}).sort("created_at", -1))
    return issues

@api.get("/issues")
async def list_all_issues(
    status: Optional[str] = None,
    priority: Optional[str] = None,
    company_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    user=Depends(require_admin_console),
):
    q: Dict[str, Any] = {}
    if status:
        q["status"] = status
    if priority:
        q["priority"] = priority
    if company_id:
        q["company_id"] = company_id
    if user.get("role") == "company_admin":
        q["company_id"] = user.get("company_id")
    elif user.get("role") == "sub_admin":
        assigned = user.get("assigned_company_ids") or []
        if assigned:
            q["company_id"] = {"$in": assigned}

    skip = (page - 1) * per_page
    total = db.ticket_issues.count_documents(q)
    items = list(db.ticket_issues.find(q, {"_id": 0}).sort("created_at", -1).skip(skip).limit(per_page))
    return {"items": items, "total": total, "page": page, "per_page": per_page}

@api.patch("/issues/{issue_id}")
async def update_issue(issue_id: str, payload: TicketIssueUpdate, user=Depends(require_admin_console)):
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    updates["updated_at"] = now_iso()
    res = db.ticket_issues.update_one({"id": issue_id}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Issue not found")
    issue = db.ticket_issues.find_one({"id": issue_id}, {"_id": 0})
    # If resolved, clear has_issue flag if no more open issues
    if payload.status in ("resolved", "cancelled"):
        open_count = db.ticket_issues.count_documents(
            {"ticket_id": issue["ticket_id"], "status": "open"}
        )
        if open_count == 0:
            db.tickets.update_one(
                {"id": issue["ticket_id"]},
                {"$set": {"has_issue": False, "updated_at": now_iso()}}
            )
    return issue

# ---------- COMPANY ADMIN MANAGEMENT (Feature 5-8) ----------
class CompanyAdminCreate(BaseModel):
    name: str
    email: EmailStr
    phone: Optional[str] = None
    password: str
    company_id: str
    designation: Optional[str] = None
    can_assign_engineers: bool = False

class CompanyAdminUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    password: Optional[str] = None
    designation: Optional[str] = None
    is_active: Optional[bool] = None
    can_assign_engineers: Optional[bool] = None

@api.get("/company-admins", dependencies=[Depends(require_admin)])
async def list_company_admins(company_id: Optional[str] = None):
    q: Dict[str, Any] = {"role": "company_admin"}
    if company_id:
        q["company_id"] = company_id
    items = list(db.users.find(q, {"_id": 0, "password_hash": 0}).sort("created_at", -1))
    return {"items": items, "total": len(items)}

@api.post("/company-admins", dependencies=[Depends(require_admin)])
async def create_company_admin(payload: CompanyAdminCreate):
    company = db.companies.find_one({"id": payload.company_id})
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")
    email = payload.email.lower().strip()
    if db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")
    doc = {
        "id": new_id(),
        "email": email,
        "name": payload.name,
        "role": "company_admin",
        "phone": payload.phone,
        "company_id": payload.company_id,
        "company_name": company["company_name"],
        "designation": payload.designation or "Company Admin",
        "can_assign_engineers": payload.can_assign_engineers,
        "password_hash": hash_password(payload.password),
        "is_active": True,
        "status": "active",
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    db.users.insert_one(doc)
    return clean({**doc})

@api.patch("/company-admins/{ca_id}", dependencies=[Depends(require_admin)])
async def update_company_admin(ca_id: str, payload: CompanyAdminUpdate):
    updates = {k: v for k, v in payload.model_dump().items() if v is not None}
    if "password" in updates:
        updates["password_hash"] = hash_password(updates.pop("password"))
    updates["updated_at"] = now_iso()
    if len(updates) == 1:
        raise HTTPException(status_code=400, detail="No changes")
    res = db.users.update_one({"id": ca_id, "role": "company_admin"}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Company admin not found")
    return clean(db.users.find_one({"id": ca_id}, {"_id": 0, "password_hash": 0}))

@api.delete("/company-admins/{ca_id}", dependencies=[Depends(require_admin)])
async def delete_company_admin(ca_id: str):
    res = db.users.delete_one({"id": ca_id, "role": "company_admin"})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Company admin not found")
    return {"ok": True}

# ---------- COMPANY ADMIN DASHBOARD (Feature 9) ----------
@api.get("/dashboard/company")
async def company_dashboard(user=Depends(get_current_user)):
    if user.get("role") not in ("admin", "sub_admin", "company_admin"):
        raise HTTPException(status_code=403, detail="Access denied")

    company_id = user.get("company_id") if user.get("role") == "company_admin" else None
    cache_k = _cache_key("dashboard-company", user)
    cached = _cache_get(cache_k)
    if cached is not None:
        return cached

    q: Dict[str, Any] = {"is_deleted": {"$ne": True}}
    if company_id:
        q["company_id"] = company_id

    counts_raw = list(db.tickets.aggregate([
        {"$match": q},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}}
    ]))
    counts = {s: 0 for s in TICKET_STATUSES}
    for item in counts_raw:
        counts[item["_id"]] = item["count"]
    counts["total"] = sum(counts.values())
    counts["active"] = sum(counts.get(s, 0) for s in [
        "open", "assigned", "accepted", "travelling",
        "reached_site", "in_progress", "resolved"
    ])
    counts["completed"] = sum(counts.get(s, 0) for s in [
        "completed_with_signature", "report_generated", "closed"
    ])

    recent_tickets = list(db.tickets.find(
        q, {"_id": 0, "id": 1, "ticket_no": 1, "ticket_number": 1,
            "customer_name": 1, "status": 1, "created_at": 1,
            "assigned_engineer_name": 1}
    ).sort("created_at", -1).limit(10))

    open_issues = db.ticket_issues.count_documents({
        "status": "open",
        **({"company_id": company_id} if company_id else {})
    })

    response = {
        "ticket_counts": counts,
        "recent_tickets": recent_tickets,
        "open_issues": open_issues,
        "company_id": company_id,
    }
    return _cache_set(cache_k, response, 30)

# ---------- COMPANY ANALYTICS (Feature 10) ----------
@api.get("/analytics/company")
async def company_analytics(user=Depends(get_current_user)):
    if user.get("role") not in ("admin", "sub_admin", "company_admin"):
        raise HTTPException(status_code=403, detail="Access denied")

    company_id = user.get("company_id") if user.get("role") == "company_admin" else None
    today = date.today()
    cache_k = _cache_key("analytics-company", user, today.isoformat())
    cached = _cache_get(cache_k)
    if cached is not None:
        return cached

    scope: Dict[str, Any] = {}
    if company_id:
        scope["company_id"] = company_id

    start_14 = (today - timedelta(days=13)).isoformat()
    per_day_raw = list(db.tickets.aggregate([
        {"$match": {"created_at": {"$gte": start_14}, **scope}},
        {"$group": {"_id": {"$substr": ["$created_at", 0, 10]}, "count": {"$sum": 1}}},
    ]))
    per_day_map = {r["_id"]: r["count"] for r in per_day_raw}
    per_day = [
        {"date": (today - timedelta(days=i)).isoformat(),
         "count": per_day_map.get((today - timedelta(days=i)).isoformat(), 0)}
        for i in range(13, -1, -1)
    ]

    status_counts_raw = list(db.tickets.aggregate([
        {"$match": scope},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}}
    ]))
    status_counts = {r["_id"]: r["count"] for r in status_counts_raw}

    closed_tickets = list(db.tickets.find(
        {"status": {"$in": ["closed", "report_generated"]},
         "completed_at": {"$ne": None}, "created_at": {"$ne": None}, **scope},
        {"_id": 0, "created_at": 1, "completed_at": 1}
    ).limit(500))

    def hours_diff2(start, end):
        try:
            from datetime import datetime as dt2
            s = dt2.fromisoformat(start[:19])
            e = dt2.fromisoformat(end[:19])
            diff = (e - s).total_seconds() / 3600
            return round(diff, 1) if diff >= 0 else None
        except Exception:
            return None

    resolution_hours = [
        h for t in closed_tickets
        if (h := hours_diff2(t.get("created_at", ""), t.get("completed_at", ""))) is not None
    ]
    avg_resolution_hours = round(sum(resolution_hours) / len(resolution_hours), 1) if resolution_hours else 0

    response = {
        "per_day": per_day,
        "status_counts": status_counts,
        "avg_resolution_hours": avg_resolution_hours,
        "total_closed": len(resolution_hours),
    }
    return _cache_set(cache_k, response, 120)

app.include_router(api)
app.add_middleware(GZipMiddleware, minimum_size=1000)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)
